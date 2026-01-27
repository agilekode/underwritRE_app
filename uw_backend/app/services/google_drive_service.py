import gspread
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from pathlib import Path
from collections import defaultdict
import pandas as pd
from flask import send_file
import io
import tempfile
import ezodf
import os
import sys
import time
import gspread.utils
from gspread.utils import rowcol_to_a1
from google.auth import default
from dotenv import load_dotenv
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl import load_workbook
import google.auth.transport.requests
import requests
from gspread.utils import a1_to_rowcol, rowcol_to_a1

load_dotenv()

# Set SSL environment variables to help with SSL issues
os.environ['PYTHONHTTPSVERIFY'] = '1'
os.environ['OPENSSL_CONF'] = ''  # Use system OpenSSL configuration

SERVICE_ACCOUNT_FILE = os.getenv("SERVICE_ACCOUNT_FILE")



SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]


PARENT_FOLDER_ID = os.getenv("GOOGLE_PARENT_FOLDER_ID")
# PARENT_FOLDER_ID = '14dBWRN-2ItiRLRy9HWfy1U3oF_p7MzPt'
# PARENT_FOLDER_ID = '1xoc6MuOW8ULr3PIucwoEhG0hwpWeIxdk'
# Determine environment: Use local creds if file is specified, otherwise use default (Cloud Run, GCP, etc.)
try:
    if SERVICE_ACCOUNT_FILE and os.path.exists("./" + SERVICE_ACCOUNT_FILE):
        print("🔑 Using local service account credentials. (gdrive)")
        print(f"📁 Service account file path: ./{SERVICE_ACCOUNT_FILE}")
        creds = Credentials.from_service_account_file("./" + SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    else:
        print("🔐 Using default GCP credentials (Cloud Run). (gdrive)")
        print(f"📁 SERVICE_ACCOUNT_FILE env var: {SERVICE_ACCOUNT_FILE}")
        creds, _ = default(scopes=SCOPES)
    
    print("✅ Credentials initialized successfully")
    
    # Configure modern HTTP transport using requests/urllib3 instead of httplib2
    import ssl
    
    # BEST PRACTICE: Create a custom HTTP transport using requests/urllib3 instead of httplib2
    import requests
    import urllib3
    import ssl
    from google.auth.transport.requests import AuthorizedSession
    
    # Disable SSL warnings for cleaner logs
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    # Create a custom HTTP transport class that bypasses httplib2 completely
    class RequestsTransport:
        def __init__(self, credentials):
            self.session = AuthorizedSession(credentials)
            # Configure session with modern SSL settings
            self.session.timeout = 60
            self.session.verify = True
            
            # Configure urllib3 with better SSL settings
            adapter = requests.adapters.HTTPAdapter(
                pool_connections=10,
                pool_maxsize=10,
                max_retries=urllib3.Retry(
                    total=3,
                    backoff_factor=0.5,
                    status_forcelist=[500, 502, 503, 504]
                )
            )
            self.session.mount('https://', adapter)
            self.session.mount('http://', adapter)
            
        def request(self, uri, method="GET", **kwargs):
            try:
                # Map httplib2 parameters to requests parameters
                requests_kwargs = {}
                
                # Handle body parameter (httplib2 uses 'body', requests uses 'data')
                if 'body' in kwargs:
                    requests_kwargs['data'] = kwargs.pop('body')
                
                # Handle headers
                if 'headers' in kwargs:
                    requests_kwargs['headers'] = kwargs.pop('headers')
                
                # Handle any other parameters
                requests_kwargs.update(kwargs)
                
                response = self.session.request(method, uri, **requests_kwargs)
                
                # Return tuple format that httplib2 expects: (response_object, content)
                # The response object needs to have a 'status' attribute
                class ResponseObject:
                    def __init__(self, status_code):
                        self.status = status_code
                        self.reason = "OK" if status_code == 200 else "Error"
                
                return ResponseObject(response.status_code), response.content
            except Exception as e:
                # Convert requests exceptions to httplib2-like format
                raise Exception(f"Request failed: {str(e)}")
                
        def close(self):
            self.session.close()
    
    # Create our custom transport
    custom_transport = RequestsTransport(creds)
    
    # Test the credentials by building the services with custom transport
    gs_client = gspread.authorize(creds)
    
    # Build services with custom transport (bypassing httplib2 completely)
    drive_service = build('drive', 'v3', http=custom_transport)
    sheets_service = build('sheets', 'v4', http=custom_transport)
    
    # Test SSL connection to Google APIs using our custom transport
    try:
        print("🔍 Testing SSL connection to Google APIs...")
        status_code, content = custom_transport.request('GET', 'https://www.googleapis.com/discovery/v1/apis')
        print(f"✅ SSL test successful: {status_code}")
    except Exception as e:
        print(f"⚠️ SSL test failed: {str(e)}")
        print("🔧 This might indicate network or SSL configuration issues")
    
    # Set environment variables for better SSL handling
    os.environ['PYTHONHTTPSVERIFY'] = '1'
    os.environ['HTTPLIB2_DEBUG'] = '0'  # Disable httplib2 debug logging
    os.environ['OPENSSL_CONF'] = ''  # Use system OpenSSL configuration
    
    # Additional SSL environment variables to help with handshake issues
    os.environ['SSL_CERT_FILE'] = ''
    os.environ['SSL_CERT_DIR'] = ''
    
    print("🔧 SSL Configuration:")
    print(f"   - OpenSSL Version: {ssl.OPENSSL_VERSION}")
    print(f"   - Python HTTPS Verify: {os.environ.get('PYTHONHTTPSVERIFY', 'Not set')}")
    print(f"   - Requests Version: {requests.__version__}")
    print(f"   - urllib3 Version: {urllib3.__version__}")
    print(f"   - Using BEST PRACTICE: requests/urllib3 transport (bypassing httplib2)")
    print(f"   - SSL errors should be eliminated with modern transport")
    
    print("✅ Google services built successfully")
    
except Exception as e:
    print(f"❌ Error initializing Google credentials: {str(e)}")
    import traceback
    traceback.print_exc()
    raise


def update_copied_sheet_values(sheet_id, mapped_values, df):
    print("🔄 Starting sheet update...")
    copied_sheet = gs_client.open_by_key(sheet_id)

    # Clean column names
    df.columns = [col.strip() for col in df.columns]
    # print(f"🧾 Cleaned columns in df: {df.columns.tolist()}")

    # Build location map: (section, field_key) → (sheet_name, cell)
    location_map = {}
    start_month_map = {}
    end_month_map = {}

    for i, row in df.iterrows():
        try:
            # print(f"🔍 Row {i} raw: {row.to_dict()}")

            section = (row.get('section') or '').strip()
            field_key = (row.get('field_key') or '').strip()
            loc_str = (row.get('location') or '').strip()
            start_month_loc_str = (row.get('start_month_location') or '').strip()
            end_month_loc_str = (row.get('end_month_location') or '').strip()

            if loc_str.startswith('=') and '!' in loc_str:
                sheet_part, cell = loc_str[1:].split('!', 1)
                location_map[(section, field_key)] = (sheet_part.strip().strip("'").strip('"'), cell.strip())

            if start_month_loc_str.startswith('=') and '!' in start_month_loc_str:
                sheet_part, cell = start_month_loc_str[1:].split('!', 1)
                start_month_map[(section, field_key)] = (sheet_part.strip().strip("'").strip('"'), cell.strip())

            if end_month_loc_str.startswith('=') and '!' in end_month_loc_str:
                sheet_part, cell = end_month_loc_str[1:].split('!', 1)
                end_month_map[(section, field_key)] = (sheet_part.strip().strip("'").strip('"'), cell.strip())

        except Exception as e:
            print(f"❌ Error processing row {i}: {e}")

    # print(f"📍 Mapped fields: main={len(location_map)}, start_month={len(start_month_map)}, end_month={len(end_month_map)}")

    # Build batch update payloads
    updates_by_sheet = defaultdict(list)

    # print("MAPPED VALUES", mapped_values)
    for item in mapped_values:
        if 'field_key' not in item:
            continue  # Only process items with 'field_key'
        key = item['field_key']
        value = item['value']
        if value == "yes":
            value = "Yes"
        if value == "no":
            value = "No"
        field_type = item.get('field_type')

        # print(f"🔍 Field type: {field_type}, Field Key: {key}, Value: {value}")

        # === Main field
        # Only use field_key for lookup (ignore section)
        # Try direct match first
        sheet_cell = None
        for map_key in location_map.keys():
            if map_key[1].strip().lower() == key.strip().lower():
                sheet_cell = location_map[map_key]
                break

        if sheet_cell:
            # print(f"🔍 Field type: {field_type}, Field Key: {key}, Value: {value}\n")
            if field_type == 'percent' and value:
                # print(f"🔍 Field type: {field_type}, Field Key: {key}, Value: {value}")
                try:
                    float(value)
                    value = f"{value}%"
                except ValueError:
                    print(f"⚠️ Could not format percent for {key}: {value}")
            sheet_name, cell = sheet_cell
            updates_by_sheet[sheet_name].append({
                'range': f"{sheet_name}!{cell}",
                'values': [[value]]
            })
            # print(f"📝 Queued update: {sheet_name}!{cell} ← {value}")
        else:
            print(f"⚠️ No location found for {key}")

        # === Optional start_month
        if 'start_month' in item:
            sm_value = item['start_month']
            sm_cell = None
            for map_key in start_month_map.keys():
                if map_key[1].strip().lower() == key.strip().lower():
                    sm_cell = start_month_map[map_key]
                    break
            if sm_cell and sm_value:
                sm_sheet, sm_loc = sm_cell
                updates_by_sheet[sm_sheet].append({
                    'range': f"{sm_sheet}!{sm_loc}",
                    'values': [[sm_value]]
                })
                print(f"📝 Queued start_month: {sm_sheet}!{sm_loc} ← {sm_value}")

        # === Optional end_month
        if 'end_month' in item:
            em_value = item['end_month']
            em_cell = None
            for map_key in end_month_map.keys():
                if map_key[1].strip().lower() == key.strip().lower():
                    em_cell = end_month_map[map_key]
                    break
            if em_cell and em_value:
                em_sheet, em_loc = em_cell
                updates_by_sheet[em_sheet].append({
                    'range': f"{em_sheet}!{em_loc}",
                    'values': [[em_value]]
                })
                print(f"📝 Queued end_month: {em_sheet}!{em_loc} ← {em_value}")

    # === Execute updates
    for sheet_name, updates in updates_by_sheet.items():
        try:
            copied_sheet.values_batch_update(
                body={
                    "valueInputOption": "USER_ENTERED",
                    "data": updates
                }
            )
            print(f"✅ Batch updated {len(updates)} cells in '{sheet_name}'")
        except Exception as e:
            print(f"❌ Failed batch update for sheet '{sheet_name}': {e}")


def extract_tables_for_storage(sheet_id, creds):
    sheets_service = build("sheets", "v4", credentials=creds)
    copied_sheet = gs_client.open_by_key(sheet_id)
    mapping_ws = copied_sheet.worksheet("Table Mapping")
    mapping_data = mapping_ws.get_all_records()

    output = []

    print(f"🔍 Processing {len(mapping_data)} table entries from mapping data")
    
    for entry in mapping_data:
        table_name = entry.get("table_name")
        location = entry.get("table_location")
        table_order = entry.get("table_order")
        
        print(f"📋 Processing table: {table_name} at location: {location}")

        if not location or '!' not in location:
            print(f"⚠️ Skipping table '{table_name}' - invalid location: {location}")
            continue

        sheet_name, range_str = location.split('!', 1)
        range_notation = f"{sheet_name}!{range_str}"

        try:
            result = sheets_service.spreadsheets().get(
                spreadsheetId=sheet_id,
                ranges=[range_notation],
                includeGridData=True
            ).execute()

            # Get the first (and only) sheet's data
            sheet_data = result['sheets'][0]['data']
            print(f"📊 API response for '{table_name}': {len(sheet_data)} data ranges")
            
            if not sheet_data:
                print(f"⚠️ No data found for table '{table_name}' at {range_notation}")
                continue
                
            grid_data = sheet_data[0].get('rowData', [])
            print(f"📋 Grid data for '{table_name}': {len(grid_data)} rows")
            values, styles = [], []

            for row in grid_data:
                row_values, row_styles = [], []

                for cell in row.get("values", []):
                    # Extract value
                    val = cell.get("formattedValue", "")
                    row_values.append(val)

                    # Extract style
                    fmt = cell.get("effectiveFormat", {})
                    text_fmt = fmt.get("textFormat", {})
                    bg_color = fmt.get("backgroundColor", {})

                    bold = text_fmt.get("bold", False)

                    def to_rgb(c): return int(c * 255)
                    r = to_rgb(bg_color.get("red", 1))
                    g = to_rgb(bg_color.get("green", 1))
                    b = to_rgb(bg_color.get("blue", 1))

                    style = f"background-color: rgb({r},{g},{b});"
                    if bold:
                        style += " font-weight: bold;"
                    # If background is dark, force white font for readability
                    # Perceived luminance formula
                    luminance = 0.299 * r + 0.587 * g + 0.114 * b
                    if luminance < 140:
                        style += " color: white;"

                    row_styles.append(style)

                values.append(row_values)
                styles.append(row_styles)

            output.append({
                "table_name": table_name,
                "table_order": table_order,
                "location": location,
                "data": values,
                "styles": styles
            })

           
            
        except Exception as e:
            print(f"❌ Failed to extract table '{table_name}' from {range_notation}: {e}")

    # Sort by table_order (handling missing or non-integer values safely)
    print("OUTPUT", output)
    output.sort(key=lambda x: x.get("table_order") if isinstance(x.get("table_order"), int) else float("inf"))
    print(f"✅ Extracted {len(output)} tables: {[t['table_name'] for t in output]}")
    return output

def extract_variables_from_sheet(sheet_id, creds):
    sheets_service = build('sheets', 'v4', credentials=creds)
    copied_sheet = gs_client.open_by_key(sheet_id)
    variable_ws = copied_sheet.worksheet("Variable Mapping")
    variable_data = variable_ws.get_all_records()

    variables = {}
    for entry in variable_data:
        name = entry.get("variable_name")
        location = entry.get("variable_location")

        if not name or not location:
            print(f"⚠️ Skipping invalid entry: {entry}")
            continue

        # Case 1: location is a formula/cell reference like =Sheet!A1
        if location.startswith('=') and '!' in location:
            try:
                location_clean = location[1:]
                sheet_name, cell = location_clean.split('!', 1)
                range_notation = f"{sheet_name}!{cell}"
                result = sheets_service.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range=range_notation,
                    valueRenderOption='FORMATTED_VALUE'
                ).execute()
                value = result.get("values", [[]])[0][0] if result.get("values") else ""
                variables[name] = value
                print(f"✅ Extracted {name} from {range_notation}: {value}")
            except Exception as e:
                print(f"⚠️ Failed to extract {name} from location {location}: {e}")

        # Case 2: location is already a value (e.g. "33.9%")
        else:
            variables[name] = location
            print(f"✅ Using literal value for {name}: {location}")

    return variables

def get_existing_folder_id(name, parent_id, max_retries=5):
    for attempt in range(max_retries):
        try:
            query = f"mimeType='application/vnd.google-apps.folder' and name='{name}' and '{parent_id}' in parents and trashed=false"
            results = drive_service.files().list(
                q=query, 
                spaces='drive', 
                fields='files(id, name, createdTime)',
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                orderBy='createdTime desc'  # Order by creation time, most recent first
            ).execute()
            files = results.get('files', [])
            
            if files:
                if len(files) > 1:
                    print(f"⚠️ Found {len(files)} folders named '{name}' in parent '{parent_id}', using most recent")
                    for i, file in enumerate(files[:3]):  # Show first 3
                        print(f"  {i+1}. ID: {file.get('id')}, Created: {file.get('createdTime')}")
                return files[0]['id']  # Return the most recent (first in ordered list)
            else:
                return None
        except Exception as e:
            # Check for SSL and network errors
            if ("ssl" in str(e).lower() or "timeout" in str(e).lower() or "timed out" in str(e).lower() or "connection" in str(e).lower()) and attempt < max_retries - 1:
                # Use longer delays for SSL errors
                if "ssl" in str(e).lower():
                    wait_time = 3 * (attempt + 1)  # 3s, 6s, 9s, 12s, 15s for SSL errors
                    error_type = "SSL"
                else:
                    wait_time = 0.5 * (attempt + 1)  # 0.5s, 1s, 1.5s for other errors
                    error_type = "network"
                print(f"⚠️ {error_type} error on attempt {attempt + 1}, retrying in {wait_time}s...")
                time.sleep(wait_time)
                continue
            else:
                raise
    return None

# === Folder Creation with Ownership Transfer ===
def create_drive_folder(name, parent_id, max_retries=5):
    print(f"🔍 Creating folder '{name}' under '{parent_id}'")
    for attempt in range(max_retries):
        try:
            existing_id = get_existing_folder_id(name, parent_id)
            print("EXISTING ID", existing_id)
            if existing_id:
                return existing_id
            file_metadata = {
                'name': name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [parent_id]
            }
            folder = drive_service.files().create(
                body=file_metadata,
                fields='id',
                supportsAllDrives=True
            ).execute()
            return folder['id']
        except Exception as e:
            # Check for SSL and network errors
            if ("ssl" in str(e).lower() or "timeout" in str(e).lower() or "timed out" in str(e).lower() or "connection" in str(e).lower()) and attempt < max_retries - 1:
                # Use longer delays for SSL errors
                if "ssl" in str(e).lower():
                    wait_time = 3 * (attempt + 1)  # 3s, 6s, 9s, 12s, 15s for SSL errors
                    error_type = "SSL"
                else:
                    wait_time = 2  # 2s for other errors
                    error_type = "network"
                print(f"⚠️ {error_type} error creating folder on attempt {attempt + 1}, retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            else:
                print(f"❌ Failed to create folder '{name}' under '{parent_id}': {e}")
                raise
    return None

def copy_template_to_folder(folder_id, new_title, template_file_id, max_retries=5):
    for attempt in range(max_retries):
        try:
            copied_file = drive_service.files().copy(
                fileId=template_file_id,
                body={
                    "name": new_title,
                    "parents": [folder_id]
                },
                supportsAllDrives=True
            ).execute()
            return copied_file['id']
        except Exception as e:
            # Check for SSL and network errors
            if ("ssl" in str(e).lower() or "timeout" in str(e).lower() or "timed out" in str(e).lower() or "connection" in str(e).lower()) and attempt < max_retries - 1:
                # Use longer delays for SSL errors
                if "ssl" in str(e).lower():
                    wait_time = 3 * (attempt + 1)  # 3s, 6s, 9s, 12s, 15s for SSL errors
                    error_type = "SSL"
                else:
                    wait_time = 2  # 2s for other errors
                    error_type = "network"
                print(f"⚠️ {error_type} error copying template on attempt {attempt + 1}, retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            else:
                print("❌ ERROR copying template:", e)
                raise  # Re-raise to trigger Flask 500 response
    return None


def generate_google_sheet_for_user_model(user_email, template_file_id):
    print("📤 Starting Google Sheet generation workflow...")
    start_time = time.time()

    try:
        # Step 1: Extract user and model info
        step_start = time.time()
        print(f"📧 User email: {user_email}")
        print(f"🆔 Template file ID: {template_file_id}")
        print(f"⏱️ Step 1 duration: {time.time() - step_start:.2f}s")

        # Step 2: Create Drive folders
        step_start = time.time()
        print("📁 Creating folders in Google Drive...")
        try:
            email_folder_id = create_drive_folder(user_email, PARENT_FOLDER_ID)
            print(f"✅ Created email folder: {email_folder_id}")
        except Exception as e:
            print(f"❌ Error creating email folder: {str(e)}")
            raise

        try:
            model_folder_name = f"Model {template_file_id}"
            model_folder_id = create_drive_folder(model_folder_name, email_folder_id)
            print(f"✅ Created model folder: {model_folder_id}")
        except Exception as e:
            print(f"❌ Error creating model folder: {str(e)}")
            raise

        print(f"✅ Created folders (Email: {email_folder_id}, Model: {model_folder_id})")
        print(f"⏱️ Step 2 duration: {time.time() - step_start:.2f}s")

        # Step 3: Copy the template
        step_start = time.time()
        current_time = time.strftime("%Y%m%d%H%M%S")
        sheet_title = f"{current_time}"
        
        try:
            copied_sheet_id = copy_template_to_folder(model_folder_id, sheet_title, template_file_id)
            print(f"✅ Successfully copied template to sheet ID: {copied_sheet_id}")
        except Exception as e:
            print(f"❌ Error copying template: {str(e)}")
            raise

        copied_sheet_url = f"https://docs.google.com/spreadsheets/d/{copied_sheet_id}"
        print(f"📄 Copied Google Sheet ID: {copied_sheet_id}")
        print(f"🔗 Sheet URL: {copied_sheet_url}")
        print(f"⏱️ Step 3 duration: {time.time() - step_start:.2f}s")

        return copied_sheet_url

    except Exception as e:
        print(f"❌ Error in generate_google_sheet_for_user_model: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


def export_google_sheet(sheet_id: str, filename: str = "exported_model.xlsx"):
    """
    Export Google Sheet as Excel file, removing internal mapping sheets.
    """

    XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    drive_export_service = build('drive', 'v3', credentials=creds)

    # Download XLSX from Google Drive
    request = drive_export_service.files().export_media(
        fileId=sheet_id,
        mimeType=XLSX_MIME
    )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
        tmp_in.write(request.execute())
        tmp_in.flush()
        tmp_in_path = tmp_in.name

    # Open and clean XLSX using openpyxl
    workbook = load_workbook(tmp_in_path)
    sheets_to_remove = {"Variable Mapping", "Table Mapping", "Model Variable Mapping"}
    sheets_to_hide = {"Underwriting Assumptions"}
    removed = []
    hidden = []

    # Remove specified sheets
    for sheet_name in sheets_to_remove:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
            removed.append(sheet_name)
    
    # Hide specified sheets
    for sheet_name in sheets_to_hide:
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].sheet_state = 'hidden'
            hidden.append(sheet_name)

    # Save cleaned XLSX to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        workbook.save(tmp_out.name)
        tmp_out_path = tmp_out.name

    with open(tmp_out_path, "rb") as f:
        xlsx_io = io.BytesIO(f.read())
    xlsx_io.seek(0)

    os.remove(tmp_in_path)
    os.remove(tmp_out_path)

    print(f"✅ XLSX file ready to send: {filename}. Removed sheets: {removed}, Hidden sheets: {hidden}")
    return send_file(
        xlsx_io,
        mimetype=XLSX_MIME,
        as_attachment=True,
        download_name=filename
    )


def extract_tables_for_storage_batch(sheet_id, table_mapping_data, sheets_service):
    range_map = {}
    ranges = []

    print("📋 Raw table mapping entries:")
    # for entry in table_mapping_data:
        # print(f"  • {entry}")
        # print(f"    - table_name: {entry.get('table_name')}")
        # print(f"    - table_location: {entry.get('table_location')}")
        # print(f"    - table_order: {entry.get('table_order')}")

    for entry in table_mapping_data:
        location = entry.get("table_location")
        if location and '!' in location:
            # Handle both formats: with or without '=' prefix
            if location.startswith('='):
                clean_location = location[1:].strip()  # Remove '=' and strip whitespace
            else:
                clean_location = location.strip()
            
            # Ensure proper sheet name formatting for Google Sheets API
            if '!' in clean_location:
                sheet_part, range_part = clean_location.split('!', 1)
                # Remove any existing quotes and re-add them properly
                sheet_name = sheet_part.strip("'\"")
                clean_location = f"'{sheet_name}'!{range_part}"
            
            print(f"✅ Valid range found: {clean_location}")
            ranges.append(clean_location)
            range_map[clean_location] = entry
        else:
            print(f"⚠️ Skipping invalid or missing location: {location}")

    if not ranges:
        print("❌ No valid ranges found. Exiting.")
        return []

    print(f"📦 Fetching {len(ranges)} table ranges from Google Sheets...")
    print(f"🔍 Final ranges to fetch: {ranges}")
    
    # First, let's check what sheets exist in the spreadsheet
    try:
        spreadsheet_info = sheets_service.spreadsheets().get(
            spreadsheetId=sheet_id,
            ranges=[],
            includeGridData=False
        ).execute()
        sheet_names = [sheet['properties']['title'] for sheet in spreadsheet_info.get('sheets', [])]
        print(f"📋 Available sheets: {sheet_names}")
    except Exception as e:
        print(f"⚠️ Could not fetch sheet info: {e}")
    
    # Make separate API calls for each range to avoid merging issues
    all_data_entries = []
    for i, range_notation in enumerate(ranges):
        try:
            print(f"🔍 Fetching range {i+1}/{len(ranges)}: {range_notation}")
            result = sheets_service.spreadsheets().get(
                spreadsheetId=sheet_id,
                ranges=[range_notation],
                includeGridData=True
            ).execute()
            
            sheet_data = result.get("sheets", [])
            if sheet_data and sheet_data[0].get("data"):
                data_entry = sheet_data[0]["data"][0]
                data_entry["range"] = range_notation  # Add range info for debugging
                all_data_entries.append(data_entry)
                print(f"✅ Successfully fetched range: {range_notation}")
            else:
                print(f"⚠️ No data returned for range: {range_notation}")
                
        except Exception as e:
            print(f"❌ Error fetching range {range_notation}: {e}")
    
    data_entries = all_data_entries

    print(f"📥 Successfully fetched {len(data_entries)} data entries")
    print(f"📊 Data entries received: {len(data_entries)}")
    print(f"📋 Ranges requested: {ranges}")
    
    # Debug each data entry
    for i, data_entry in enumerate(data_entries):
        print(f"📄 Data entry {i}: {data_entry.get('range', 'No range info')}")
        print(f"   Rows: {len(data_entry.get('rowData', []))}")
    
    if len(data_entries) != len(ranges):
        print(f"⚠️ Mismatch: expected {len(ranges)} ranges but got {len(data_entries)} grid data blocks")
        print(f"🔍 This might indicate that some ranges are invalid or overlapping")
        print(f"🔍 Missing ranges: {set(ranges) - set([entry.get('range', '') for entry in data_entries])}")

    output = []

    for i, grid in enumerate(data_entries):
        try:
            grid_data = grid.get('rowData', [])
            location = ranges[i]
            table_entry = range_map.get(location)
            table_name = table_entry.get("table_name")
            table_order = table_entry.get("table_order")
            summary = table_entry.get("summary", "FALSE")

            print(f"🔍 Processing table: {table_name} from {location}")

            values, styles = [], []

            for row in grid_data:
                row_values, row_styles = [], []

                for cell in row.get("values", []):
                    val = cell.get("formattedValue", "")
                    row_values.append(val)

                    fmt = cell.get("effectiveFormat", {})
                    text_fmt = fmt.get("textFormat", {})
                    bg_color = fmt.get("backgroundColor", {})

                    r = int(bg_color.get("red", 1) * 255)
                    g = int(bg_color.get("green", 1) * 255)
                    b = int(bg_color.get("blue", 1) * 255)
                    style = f"background-color: rgb({r},{g},{b});"
                    if text_fmt.get("bold", False):
                        style += " font-weight: bold;"
                    luminance = 0.299 * r + 0.587 * g + 0.114 * b
                    if luminance < 140:
                        style += " color: white;"

                    row_styles.append(style)

                values.append(row_values)
                styles.append(row_styles)

            output.append({
                "table_name": table_name,
                "table_order": table_order,
                "summary": summary,
                "location": location,
                "data": values,
                "styles": styles
            })

        except Exception as e:
            print(f"❌ Error processing range {ranges[i]}: {e}")

    print(f"✅ Successfully extracted {len(output)} tables")
    output.sort(key=lambda x: x.get("table_order") if isinstance(x.get("table_order"), int) else float("inf"))
    return output

def extract_variables_from_sheet_batch(sheet_id, variable_data, sheets_service):
    """
    Extracts variables from a Google Sheet in batch, minimizing API calls for speed.
    - If a variable location is a literal, it's used directly.
    - If a variable location is a formula (starts with '=' and contains '!'), it's batched for a single API call.
    """
    time.sleep(1)
    variables = {}
    ranges = []
    location_map = {}

    # Preprocess: collect all ranges and literals in one pass
    for entry in variable_data:
        name = entry.get("variable_name")
        location = entry.get("variable_location")
        if not name or not location:
            continue
        if location.startswith('=') and '!' in location:
            clean_location = location[1:]  # Remove '='
            ranges.append(clean_location)
            location_map[clean_location] = name
        else:
            variables[name] = location  # Literal value

    # Batch get all ranges in a single API call (already optimal)
    if ranges:
        # Google Sheets API allows up to 100 ranges per batchGet call.
        # If more, split into chunks to avoid multiple roundtrips.
        CHUNK_SIZE = 100
        for i in range(0, len(ranges), CHUNK_SIZE):
            chunk = ranges[i:i+CHUNK_SIZE]
            result = sheets_service.spreadsheets().values().batchGet(
                spreadsheetId=sheet_id,
                ranges=chunk,
                valueRenderOption='FORMATTED_VALUE'
            ).execute()
            for j, value_range in enumerate(result.get("valueRanges", [])):
                name = location_map[chunk[j]]
                values = value_range.get("values", [[]])
                value = values[0][0] if values and values[0] else ""
                variables[name] = value

    return variables

def get_market_rent_insert_ops(market_ws, market_json, rental_assumptions_json, sheet_name="Market Rent Assumptions", rental_sheet_name="Rental Assumptions"):
    market_start_row = 2
    market_num_rows = len(market_json)
    market_end_row = market_start_row + market_num_rows - 1

    rental_start_row = 2
    rental_end_row = rental_start_row + len(rental_assumptions_json) - 1

    rental_d_range = f"'{rental_sheet_name}'!$D${rental_start_row}:$D${rental_end_row}"
    rental_h_range = f"'{rental_sheet_name}'!$H${rental_start_row}:$H${rental_end_row}"
    # Additional ranges for filters:
    # C: vacate_flag; F: vacate_month
    rental_c_range = f"'{rental_sheet_name}'!$C${rental_start_row}:$C${rental_end_row}"
    rental_f_range = f"'{rental_sheet_name}'!$F${rental_start_row}:$F${rental_end_row}"

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": market_ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": market_start_row - 1,
                "endIndex": market_start_row - 1 + market_num_rows
            },
            "inheritFromBefore": False
        }
    }

    values = []
    for i, entry in enumerate(market_json):
        row_num = market_start_row + i
        # Average of current rents (H) for matching layout (D == A{row_num}),
        # but only include rows where: C = 0 OR ((C = 1 OR C = 2) AND F > 0)
        # Using FILTER with boolean arithmetic for OR/AND logic
        formula = (
            f"=IFERROR("
            f"SUMPRODUCT({rental_h_range}*--({rental_d_range}=A{row_num})*("
            f"--({rental_c_range}=0)+((--({rental_c_range}=1)+--({rental_c_range}=2))*--({rental_f_range}>0))"
            f"))/"
            f"SUMPRODUCT(--({rental_d_range}=A{row_num})*("
            f"--({rental_c_range}=0)+((--({rental_c_range}=1)+--({rental_c_range}=2))*--({rental_f_range}>0))"
            f")),0)"
        )
        values.append([entry["layout"], "", entry["pf_rent"], formula])

    value_data = {
        "range": f"'{sheet_name}'!A{market_start_row}:D{market_end_row}",
        "values": values
    }

    # Format request: color Column C text blue for the inserted rows
    # Column C is 0-based index 2, end index is non-inclusive 3
    format_request = {
        "repeatCell": {
            "range": {
                "sheetId": market_ws._properties["sheetId"],
                "startRowIndex": market_start_row - 1,
                "endRowIndex": market_start_row - 1 + market_num_rows,
                "startColumnIndex": 2,
                "endColumnIndex": 3
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {
                        "foregroundColor": {"red": 0, "green": 0, "blue": 1}
                    }
                }
            },
            "fields": "userEnteredFormat.textFormat.foregroundColor"
        }
    }

    return insert_request, value_data, format_request

def get_rental_assumptions_insert_ops(rental_ws, rental_assumptions_json, market_start_row, market_end_row, sheet_name="Rental Assumptions"):
    rental_start_row = 2
    rental_num_rows = len(rental_assumptions_json)
    rental_end_row = rental_start_row + rental_num_rows - 1
    total_row_index = rental_end_row + 1

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": rental_ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": rental_start_row - 1,
                "endIndex": rental_start_row - 1 + rental_num_rows
            },
            "inheritFromBefore": False
        }
    }

  
    format_request = [
        # Blue text for columns A–F (0-based: 0–6)
        {
            "repeatCell": {
                "range": {
                    "sheetId": rental_ws._properties["sheetId"],
                    "startRowIndex": rental_start_row - 1,
                    "endRowIndex": rental_start_row - 1 + rental_num_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": 6  # A=0, B=1, C=2, D=3, E=4, F=5
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {
                            "foregroundColor": {"red": 0, "green": 0, "blue": 1}
                        }
                    }
                },
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
                {
            "repeatCell": {
                "range": {
                    "sheetId": rental_ws._properties["sheetId"],
                    "startRowIndex": rental_start_row - 1,
                    "endRowIndex": rental_start_row - 1 + rental_num_rows,
                    "startColumnIndex": 7,
                    "endColumnIndex": 8  # G=6, H=7, I=8
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {
                            "foregroundColor": {"red": 0, "green": 0, "blue": 1}
                        }
                    }
                },
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        # Date format for column G (vacate_month, 0-based index 6)
        {
            "repeatCell": {
                "range": {
                    "sheetId": rental_ws._properties["sheetId"],
                    "startRowIndex": rental_start_row - 1,
                    "endRowIndex": rental_start_row - 1 + rental_num_rows,
                    "startColumnIndex": 6,
                    "endColumnIndex": 7
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "DATE",
                            "pattern": "m/dd/yy"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        }
    ]

    values = []
    for i, item in enumerate(rental_assumptions_json):
        row_num = rental_start_row + i
        pf_rent_formula = (
            f"=IFERROR("
            f"IF(OR(C{row_num}=1,C{row_num}=2),"
            f"INDEX('Market Rent Assumptions'!$C${market_start_row}:$C${market_end_row}, "
            f"MATCH(D{row_num}, 'Market Rent Assumptions'!$A${market_start_row}:$A${market_end_row}, 0)),"
            f"H{row_num}),"
            f"H{row_num})"
        )
        annual_pf_rent_formula = f"=I{row_num}*12"
        values.append([
            i + 1,
            item["rent_type"],
            item["vacate_flag"],
            item["layout"],
            item["square_feet"],
            item["vacate_month"],
            f"=EOMONTH(Assumptions!$F$2,F{row_num})",
            item["current_rent"],
            pf_rent_formula,
            annual_pf_rent_formula
        ])

    value_data = {
        "range": f"'{sheet_name}'!A{rental_start_row}:J{rental_end_row}",
        "values": values
    }

    total_row_data = {
        "range": f"'{sheet_name}'!A{total_row_index}:J{total_row_index}",
        "values": [[
            "", "", "", 
            f"=COUNTA(A{rental_start_row}:A{rental_end_row})",
            f"=SUM(E{rental_start_row}:E{rental_end_row})/D{total_row_index}",
            "", "",
            f"=SUM(H{rental_start_row}:H{rental_end_row})",
            f"=SUM(I{rental_start_row}:I{rental_end_row})",
            f"=SUM(J{rental_start_row}:J{rental_end_row})"
        ]]
    }

    return [insert_request, *format_request], value_data, total_row_data

def get_market_rent_formula_updates(market_json, rental_start_row, rental_end_row, market_start_row, sheet_name="Market Rent Assumptions", rental_sheet_name="Rental Assumptions"):
    # Define ranges used in formulas
    rental_d_range = f"'{rental_sheet_name}'!$D${rental_start_row}:$D${rental_end_row}"
    rental_e_range = f"'{rental_sheet_name}'!$E${rental_start_row}:$E${rental_end_row}"
    rental_h_range = f"'{rental_sheet_name}'!$H${rental_start_row}:$H${rental_end_row}"
    rental_i_range = f"'{rental_sheet_name}'!$I${rental_start_row}:$I${rental_end_row}"
    # Additional filter ranges for eligibility:
    # C: vacate_flag; F: vacate_month
    rental_c_range = f"'{rental_sheet_name}'!$C${rental_start_row}:$C${rental_end_row}"
    rental_f_range = f"'{rental_sheet_name}'!$F${rental_start_row}:$F${rental_end_row}"

    # Build formulas row by row for columns D–G
    avg_current_rent = []
    for i in range(len(market_json)):
        row_index = market_start_row + i
        row_formulas = [
            # Average current rent (H) for matching layout (D==A{row_index}) including only rows where:
            # C = 0 OR ((C = 1 OR C = 2) AND F > 0)
            f"=IFERROR("
            f"SUMPRODUCT({rental_h_range}*--({rental_d_range}=A{row_index})*("
            f"--({rental_c_range}=0)+((--({rental_c_range}=1)+--({rental_c_range}=2))*--({rental_f_range}>0))"
            f"))/"
            f"SUMPRODUCT(--({rental_d_range}=A{row_index})*("
            f"--({rental_c_range}=0)+((--({rental_c_range}=1)+--({rental_c_range}=2))*--({rental_f_range}>0))"
            f")),0)",  # D
            f"=IFERROR(SUMIF({rental_d_range},A{row_index},{rental_h_range})/SUMIF({rental_d_range},A{row_index},{rental_e_range}),0)",  # E
            f"=IFERROR(AVERAGEIF({rental_d_range},A{row_index},{rental_i_range}),0)",  # F
            f"=IFERROR(SUMIF({rental_d_range},A{row_index},{rental_i_range})/SUMIF({rental_d_range},A{row_index},{rental_e_range}),0)"   # G
        ]
        avg_current_rent.append(row_formulas)

    # Return batch update payload
    return {
        "range": f"'{sheet_name}'!D{market_start_row}:G{market_start_row + len(market_json) - 1}",
        "values": avg_current_rent
    }



import re

def a1_to_row_col(cell):
    """Convert A1 notation (e.g., 'B12' or 'Sheet1!B12') to 0-based row, col."""
    if '!' in cell:
        _, cell = cell.split('!', 1)  # Remove sheet name

    match = re.match(r"^([A-Z]+)(\d+)$", cell)
    if not match:
        raise ValueError(f"Invalid A1 cell reference: {cell}")
    
    col_letters, row = match.groups()

    # Convert column letters to number (A=0, B=1, ..., Z=25, AA=26, etc.)
    col = 0
    for c in col_letters:
        col = col * 26 + (ord(c.upper()) - ord('A') + 1)
    return int(row) - 1, col - 1



def get_rental_growth_assumptions_inserts(assumptions_ws, rental_growth_json, model_variable_mapping):
    rental_growth_json = sorted(rental_growth_json, key=lambda x: x.get("type") == "rental")
    header_cell_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Growth Rates Header')
    if not header_cell_location:
        raise ValueError("Growth Rates Header location not found.")

    start_row, start_col = a1_to_row_col(header_cell_location)
    start_row += 1  # one row below header

    num_rows = len(rental_growth_json)
    end_row = start_row + num_rows

    # Column letters for name/value
    col_letter_name = chr(ord('A') + start_col)
    col_letter_value = chr(ord('A') + start_col + 1)

    # Prepare values
    values = [
        [entry["name"], entry["value"] / 100]
        for entry in rental_growth_json
    ]

    # === Row + column mapping: {name: {"row": X, "name_col": 'M', "value_col": 'N'}}
    row_mapping = {
        entry["name"]: {
            "row": start_row + i + 1,  # 1-indexed row
            "name_col": col_letter_name,
            "value_col": col_letter_value
        }
        for i, entry in enumerate(rental_growth_json)
    }

    value_range = f"'Assumptions'!{col_letter_name}{start_row+1}:{col_letter_value}{end_row}"
    value_payload = {
        "range": value_range,
        "values": values
    }

    format_request = {
        "repeatCell": {
            "range": {
                "sheetId": assumptions_ws._properties["sheetId"],
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col + 1,
                "endColumnIndex": start_col + 2
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "PERCENT",
                        "pattern": "0.0%"
                    },
                    "textFormat": {
                        "foregroundColor": {"red": 0, "green": 0, "blue": 1}
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.textFormat.foregroundColor"
        }
    }

    return value_payload, format_request, row_mapping

def get_rent_roll_growth_inserts(rent_ws, rental_growth_json, assumptions_row_mapping, rent_start_row=3):
    rental_entries = [entry for entry in rental_growth_json if entry.get("type") == "rental"]
    num_rows = len(rental_entries)
    end_row = rent_start_row + num_rows - 1

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": rent_ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": rent_start_row - 1,  # Insert at row 3 (0-based index 2)
                "endIndex": rent_start_row - 1 + num_rows  # End after inserting num_rows
            },
            "inheritFromBefore": False
        }
    }

    rows = []
    for entry in rental_entries:
        mapping = assumptions_row_mapping.get(entry["name"])
        if not mapping:
            raise ValueError(f"No row mapping found for rental growth entry: {entry['name']}")

        row = [""] * 10
        row[1] = f"='Assumptions'!{mapping['name_col']}{mapping['row']}"   # Column B
        row[3] = f"='Assumptions'!{mapping['value_col']}{mapping['row']}"  # Column D
        row[9] = 1
        rows.append(row)

    value_payload = {
        "range": f"'Rent Roll Model'!A{rent_start_row}:J{end_row}",
        "values": rows
    }

    return insert_request, value_payload

def get_noi_growth_factor_formula_update(
    assumption_row_mapping,
    rental_growth_json,
    noi_formula_row=9,
    noi_formula_col_start=5,
    noi_formula_col_end=140,
    sheet_name="NOI Walk"
):
    # Get rental entries and their (value_col, row) from the mapping
    rental_cells = [
        (mapping["value_col"], mapping["row"])
        for entry in rental_growth_json
        if entry.get("type") == "rental" and (mapping := assumption_row_mapping.get(entry["name"]))
    ]

    if not rental_cells:
        raise ValueError("No rental entries with value_col found in assumption_row_mapping.")

    # Build MAX range expression across all rental cells (e.g., 'Assumptions'!N51, 'Assumptions'!N52, ...)
    max_inputs = [f"'Assumptions'!{col}{row}" for col, row in rental_cells]
    max_expr = f"MAX({','.join(max_inputs)})"

    # Build the row of formulas
    row_values = []
    for col in range(noi_formula_col_start, noi_formula_col_end + 1):
        col_letter = rowcol_to_a1(1, col).replace("1", "")
        formula = f"=(1 + ({max_expr} / 12))^{col_letter}15"
        row_values.append(formula)

    range_str = f"'{sheet_name}'!{rowcol_to_a1(noi_formula_row, noi_formula_col_start)}:{rowcol_to_a1(noi_formula_row, noi_formula_col_end)}"

    return {
        "range": range_str,
        "values": [row_values]
    }


def get_rent_roll_growth_formula_updates(rental_growth_ws, rental_growth_json, growth_start_row=3, start_col_index=10, num_columns=131):

    rental_growth_json = [entry for entry in rental_growth_json if entry.get("type") == "rental"]
    growth_num_rows = len(rental_growth_json)
    formulas = []

    for i in range(growth_num_rows):
        row_num = growth_start_row + i
        row_formulas = []
        for j in range(num_columns):
            col_index = start_col_index + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            header_row = 6 + growth_num_rows  # e.g., 9 when growth_num_rows == 3
            formula = f"=(1+($D{row_num}/12))^{col_letter}${header_row}"
            row_formulas.append(formula)
        formulas.append(row_formulas)

    start_col_letter = rowcol_to_a1(1, start_col_index).replace("1", "")
    end_col_index = start_col_index + num_columns - 1
    end_col_letter = rowcol_to_a1(1, end_col_index).replace("1", "")

    value_payload = {
        "range": f"'Rent Roll Model'!{start_col_letter}{growth_start_row}:{end_col_letter}{growth_start_row + growth_num_rows - 1}",
        "values": formulas
    }

    format_request = {
        "repeatCell": {
            "range": {
                "sheetId": rental_growth_ws._properties["sheetId"],
                "startRowIndex": growth_start_row - 1,
                "endRowIndex": growth_start_row - 1 + growth_num_rows,
                "startColumnIndex": start_col_index - 1,
                "endColumnIndex": end_col_index
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "NUMBER",
                        "pattern": "0.000"
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat"
        }
    }

    return value_payload, format_request



def get_rent_roll_assumption_row_inserts(rent_roll_ws, rental_assumptions_json, rental_growth_json, model_variable_mapping, start_row=7):


    rental_growth_json = [entry for entry in rental_growth_json if entry.get("type") == "rental"]

    num_rental_rows = len(rental_assumptions_json)
    rental_rows = 2
    growth_rows = rental_growth_json
    growth_length = len(growth_rows)
    start_row = start_row + growth_length

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": rent_roll_ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": start_row -1,
                "endIndex": start_row - 1 + num_rental_rows
            },
            "inheritFromBefore": False
        }
    }
    print("RENT ROLL assumptions", rental_assumptions_json)

    print("RENT ROLL insert_request", insert_request)
    rows_to_insert = []


    rehab_time_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Rehab Time')
    lease_up_time_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Lease-up Time')
    
    for i in range(num_rental_rows):
        rental_row = rental_rows + i
        current_row = start_row + i

        row = [
            "",  # Column A
            f"='Rental Assumptions'!A{rental_row}",  # B
            f"='Rental Assumptions'!B{rental_row}",  # C
            f"='Rental Assumptions'!C{rental_row}",  # D
            f"='Rental Assumptions'!F{rental_row}",  # E
            f"=IF(D{current_row}=1,E{current_row}+{rehab_time_location}+{lease_up_time_location},E{current_row})",  # F
            f"='Rental Assumptions'!H{rental_row}",  # G
            f"='Rental Assumptions'!I{rental_row}",  # H
            ""
        ]

        month_row = 6 + growth_length

        for col_idx in range(10, 142):  # Columns J to EU
            col_letter = gspread.utils.rowcol_to_a1(1, col_idx).replace("1", "")
            formula = (
                f"=IFERROR(IF($D{current_row}=0,$G{current_row},"
                f"IF({col_letter}${month_row}<$E{current_row},$G{current_row},"
                f"IF({col_letter}${month_row}>=$F{current_row},$H{current_row},0)))"
                f"*INDEX({col_letter}$3:{col_letter}${3 + growth_length - 1},MATCH($C{current_row},$B$3:$B${3 + growth_length - 1},0)),0)"
            )
            row.append(formula)

        rows_to_insert.append(row)

    num_columns = max(len(r) for r in rows_to_insert) - 1
    end_col_letter = gspread.utils.rowcol_to_a1(1, num_columns + 1).replace("1", "")

    value_payload = {
        "range": f"'Rent Roll Model'!B{start_row}:{end_col_letter}{start_row + num_rental_rows - 1}",
        "values": [row[1:] for row in rows_to_insert]  # skip column A
    }

    return insert_request, value_payload


def get_total_summary_row_updates(rent_roll_ws, sheet_name, start_row, num_rows, model_variable_mapping, start_col_index=10, num_columns=132, ):
    # === Row math
    sum_row = start_row + num_rows
    logic_start_row = sum_row + 1
    weighted_row = sum_row + 4

    start_col_letter = rowcol_to_a1(1, start_col_index).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col_index + num_columns - 1).replace("1", "")

    # === SUM row formulas
    sum_values = [""] * (start_col_index - 2) + ["Rent Roll"]
    for col_idx in range(start_col_index, start_col_index + num_columns):
        col_letter = rowcol_to_a1(1, col_idx).replace("1", "")
        sum_values.append(f"=SUM({col_letter}{start_row}:{col_letter}{start_row + num_rows - 1})")

    # === Logic rows (3 rows)
    logic_rows = []
    for r in range(3):
        row_values = [""] * (start_col_index - 1)
        for col_idx in range(start_col_index, start_col_index + num_columns):
            col_letter = rowcol_to_a1(1, col_idx).replace("1", "")
            data_range = f"{col_letter}{start_row}:{col_letter}{start_row + num_rows - 1}"
            if r == 0:
                formula = f"=COUNTIF({data_range},\">0\")"
            elif r == 1:
                formula = f"=-COUNTIF({data_range},0)"
            elif r == 2:
                formula = (
                    f"=COUNTIFS($F${start_row}:$F${start_row + num_rows - 1},{col_letter}{start_row - 1},"
                    f"$D${start_row}:$D${start_row + num_rows - 1},1)"
                )
            row_values.append(formula)
        logic_rows.append(row_values)

    # === Reference values (F3:F4, H3:H4)
    re_stabilization_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Re-Stabilization Occurs')
    vacancy_location = get_mapped_cell_location(model_variable_mapping, 'Leasing Assumptions', 'Vacancy')

    ref_values = [
        ["Re-stabilization Occurs", "", f"={re_stabilization_location}"],
        ["Vacancy", "", f"={vacancy_location}"]
    ]

    # === Weighted row
    weighted_values = [""] * (start_col_index - 1)
    for col_idx in range(start_col_index, start_col_index + num_columns):
        col_letter = rowcol_to_a1(1, col_idx).replace("1", "")
        row_range = f"{col_letter}{start_row}:{col_letter}{start_row + num_rows - 1}"
        total_cell = f"{col_letter}{sum_row}"
        # Vacancy SUM row index = 15 + rental_growth_len + num_units*2
        # Given start_row = 7 + rental_growth_len, derive as: start_row + 8 + 2*num_rows
        vacancy_sum_row_index = start_row + 7 + (2 * num_rows)
        formula = (
            f"=-IF({col_letter}{start_row - 1}<$H$3,"
            f"{col_letter}{vacancy_sum_row_index}-{total_cell},"
            f"{total_cell}*$H$4)"
        )
        weighted_values.append(formula)

    # === Build update payloads

    update_payload = [
        {
            "range": f"'{sheet_name}'!{'A'}{sum_row}:{end_col_letter}{sum_row}",
            "values": [sum_values[0:]]
        },
        {
            "range": f"'{sheet_name}'!{start_col_letter}{logic_start_row}:{end_col_letter}{logic_start_row + 2}",
            "values": [r[start_col_index - 1:] for r in logic_rows]
        },
        {
            "range": f"'{sheet_name}'!F3:H4",
            "values": ref_values
        },
        {
            "range": f"'{sheet_name}'!{start_col_letter}{weighted_row}:{end_col_letter}{weighted_row}",
            "values": [weighted_values[start_col_index - 1:]]
        }
    ]

    format_payload = {
        "updateBorders": {
            "range": {
                "sheetId": rent_roll_ws._properties["sheetId"],  # You need to pass this into the function or access it from assumptions_ws._properties["sheetId"]
                "startRowIndex": sum_row - 1,
                "endRowIndex": sum_row,
                "startColumnIndex": 1,
                "endColumnIndex": start_col_index + num_columns - 1
            },
            "top": {
                "style": "SOLID",
                "width": 1,
                "color": {"red": 0, "green": 0, "blue": 0}
            }
        }
    }

    return update_payload, format_payload

def get_amenity_income_update_payload(sheet_name, amenity_income_json, start_row=2):
    num_rows = len(amenity_income_json)
    rows_to_insert = []

    for i, item in enumerate(amenity_income_json):
        row_num = start_row + i
        row = [
            item["name"],                            # A
            "",                                      # B
            item["start_month"],                     # C
            f"{item['utilization']}%",                 # D
            item["unit_count"],                      # E
            f"=ROUND(E{row_num}*D{row_num},0)",      # F
            item["monthly_fee"],                     # G
            f"=G{row_num}*F{row_num}",               # H
            f"=H{row_num}*12"                         # I
        ]
        rows_to_insert.append(row)

    range_str = f"'{sheet_name}'!A{start_row}:I{start_row + num_rows - 1}"

    return {
        "range": range_str,
        "values": rows_to_insert
    }


def get_amenity_income_format_requests(spreadsheet, sheet_name, start_row=2, num_amenities=4):
    """
    Returns a list of format requests for the Amenity Income sheet:
    - Columns E and F: "<value> units" (unless 0, then '-')
    - Columns G, H, I: "$<value>" (unless 0, then '-')
    """

    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    # Google Sheets column indices are 0-based
    # E = 4, F = 5, G = 6, H = 7, I = 8
    end_row = start_row + num_amenities

    # Format for "units" columns (E, F)
    units_format = {
        "numberFormat": {
            "type": "NUMBER",
            "pattern": '[=0]"-";[=1]#,##0" unit";#,##0" units"'
        }
    }
    # Format for "$" columns (G, H, I)
    dollar_format = {
        "numberFormat": {
            "type": "NUMBER",
            "pattern": '[=0]"-";"$"#,##0'
        }
    }

    # Get sheetId from the worksheet name (must be provided by caller)
    # Here, we assume the caller will replace <SHEET_ID> with the actual id
    # or you can pass the worksheet object instead of name
    format_requests = [
        # E (units)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,  # To be filled in by caller if needed
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 4,
                    "endColumnIndex": 5
                },
                "cell": {
                    "userEnteredFormat": units_format
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        },
        # F (units)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 5,
                    "endColumnIndex": 6
                },
                "cell": {
                    "userEnteredFormat": units_format
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        },
        # G (dollar)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 6,
                    "endColumnIndex": 7
                },
                "cell": {
                    "userEnteredFormat": dollar_format
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        },
        # H (dollar)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 7,
                    "endColumnIndex": 8
                },
                "cell": {
                    "userEnteredFormat": dollar_format
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        },
        # I (dollar)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,
                    "endColumnIndex": 9
                },
                "cell": {
                    "userEnteredFormat": dollar_format
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": start_row - 1 + num_amenities,
                    "startColumnIndex": 2,
                    "endColumnIndex": 5  # C, D, E (2, 3, 4)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": start_row - 1 + num_amenities,
                    "startColumnIndex": 6,
                    "endColumnIndex": 7
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    # Ensure all amenity rows are not bold (regular text)
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": end_row,
                "startColumnIndex": 0,   # from column A
                "endColumnIndex": 200    # up to a wide column bound
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {
                        "bold": False
                    }
                }
            },
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })

    # Keep the totals row bold (it's the row immediately after the amenity rows)
    # With one amenity and start_row=2, total row is 3 → 0-based index = end_row - 1
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": max((end_row - 1), 0),
                "endRowIndex": end_row,
                "startColumnIndex": 0,
                "endColumnIndex": 200
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {
                        "bold": True
                    }
                }
            },
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })

    return format_requests 




def get_amenity_income_totals_update_payload(sheet_name, start_row=2, num_amenities=4):
    total_row = start_row + num_amenities  # One row below last amenity row

    h_sum_formula = f"=SUM(H{start_row}:H{total_row - 1})"
    i_sum_formula = f"=SUM(I{start_row}:I{total_row - 1})"

    # Build row with blank columns up to G (index 6), then formulas for H and I
    total_row_values = ["Total Amenity Income"] + [""] * 6 + [h_sum_formula, i_sum_formula]

    # A1 range for entire row A–I
    range_str = f"'{sheet_name}'!A{total_row}:I{total_row}"

    return {
        "range": range_str,
        "values": [total_row_values]
    }


def get_amenity_income_insert_request(
    spreadsheet, amenity_income_json, start_row=2, sheet_name="Amenity Income"
):
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    num_amenities = len(amenity_income_json)

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_amenities,
            },
            "inheritFromBefore": False
        }
    }

    return insert_request

def get_amenity_income_insert_request_noi(
    spreadsheet, amenity_income_json, start_row=25, sheet_name="NOI Walk"
):
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    num_amenities = len(amenity_income_json)

    return {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_amenities,
            },
            "inheritFromBefore": True
        }
    }

def get_amenity_income_formula_update(
    amenity_json,
    noi_sheet="NOI Walk",
    amenity_sheet="Amenity Income",
    start_row=25,
    start_col=5,
    num_months=132
):
    print("get_amenity_income_formula_update")
    print("amenity_json", amenity_json)
    num_amenities = len(amenity_json)
    all_rows = []

    for i in range(num_amenities):
        amenity_row = 2 + i
        noi_row = start_row + i

        row = [""]  # A
        row.append(f"='{amenity_sheet}'!A{amenity_row}")  # B
        row += ["", ""]  # C, D, E padding up to start_col=4

        for j in range(num_months):
            col_index = start_col + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            formula = (
                f"=IF({col_letter}$15<'{amenity_sheet}'!$C{amenity_row},"
                f"0,"
                f"'{amenity_sheet}'!G{amenity_row}*{col_letter}$16*{col_letter}$10*'{amenity_sheet}'!D{amenity_row})"
            )
            row.append(formula)

        all_rows.append(row)

    total_cols = len(all_rows[0])
    end_col_letter = rowcol_to_a1(1, total_cols).replace("1", "")
    end_row = start_row + num_amenities - 1
    range_str = f"'{noi_sheet}'!A{start_row}:{end_col_letter}{end_row}"

    return {
        "range": range_str,
        "values": all_rows
    }



def get_amenity_income_formula_update_industrial(
    amenity_json,
    noi_sheet="NOI Walk",
    amenity_sheet="Amenity Income",
    start_row=25,
    start_col=5,
    num_months=132
):
    print("get_amenity_income_formula_update")
    print("amenity_json", amenity_json)
    num_amenities = len(amenity_json)
    all_rows = []

    for i in range(num_amenities):
        amenity_row = 2 + i
        noi_row = start_row + i

        row = [""]  # A
        row.append(f"='{amenity_sheet}'!A{amenity_row}")  # B
        row += ["", ""]  # C, D, E padding up to start_col=4

        for j in range(num_months):
            col_index = start_col + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            formula = (
                f"=IF({col_letter}$15<'{amenity_sheet}'!$C{amenity_row},"
                f"0,"
                f"'{amenity_sheet}'!G{amenity_row}*{col_letter}$10*'{amenity_sheet}'!D{amenity_row})"
            )
            row.append(formula)

        all_rows.append(row)

    total_cols = len(all_rows[0])
    end_col_letter = rowcol_to_a1(1, total_cols).replace("1", "")
    end_row = start_row + num_amenities - 1
    range_str = f"'{noi_sheet}'!A{start_row}:{end_col_letter}{end_row}"

    return {
        "range": range_str,
        "values": all_rows
    }



def get_noi_summary_insert_request(
    spreadsheet, num_rows, start_row=10, sheet_name="NOI"
):
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]

    return {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_rows,
            },
            "inheritFromBefore": True
        }
    }




def get_noi_summary_row_update_payload(
    num_rows,
    walk_start_row=25,
    noi_start_row=10,
    noi_sheet="NOI",
    walk_sheet="NOI Walk"
):
    all_rows = []

    for i in range(num_rows):
        walk_row = walk_start_row + i
        row = [""] * 2  # Columns A and B
        row.append(f"='{walk_sheet}'!B{walk_row}")  # Column C pulls name from Walk sheet
        row += ["", "", ""]  # Columns D, E, F

        for col_letter in ["G", "H", "I", "J", "K"]:
            formula = (
                f"=SUMIFS('{walk_sheet}'!$E{walk_row}:$CW{walk_row},"
                f"'{walk_sheet}'!$E$14:$CW$14,"
                f"{noi_sheet}!{col_letter}$4)"
            )
            row.append(formula)

        all_rows.append(row)

    # Define range to update
    total_cols = len(all_rows[0])
    end_col_letter = rowcol_to_a1(1, total_cols).replace("1", "")
    end_row = noi_start_row + num_rows - 1
    range_str = f"'{noi_sheet}'!A{noi_start_row}:{end_col_letter}{end_row}"

    return {
        "range": range_str,
        "values": all_rows
    }

def get_sheet_id(spreadsheet, sheet_name):
    """
    Returns the numeric sheet ID for a given sheet name in the spreadsheet.
    """
    for sheet in spreadsheet.fetch_sheet_metadata()["sheets"]:
        if sheet["properties"]["title"] == sheet_name:
            return sheet["properties"]["sheetId"]
    raise ValueError(f"Sheet '{sheet_name}' not found in spreadsheet.")

def get_noi_summary_row_insert_request(spreadsheet, sheet_name, insert_start_row, num_rows):
    return {
        "insertRange": {
            "range": {
                "sheetId": get_sheet_id(spreadsheet, sheet_name),
                "startRowIndex": insert_start_row - 1,
                "endRowIndex": insert_start_row - 1 + num_rows
            },
            "shiftDimension": "ROWS"
        }
    }

def get_operating_expenses_insert_request(spreadsheet, expenses_json, start_row=2, sheet_name="Operating Expenses"):
    sheet_id = get_sheet_id(spreadsheet, sheet_name)
    num_rows = len(expenses_json)

    return {
        "insertRange": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": start_row - 1 + num_rows
            },
            "shiftDimension": "ROWS"
        }
    }


def get_operating_expenses_insert_request(spreadsheet, expenses_json, start_row=2, sheet_name="Operating Expenses"):
    sheet_id = get_sheet_id(spreadsheet, sheet_name)
    num_rows = len(expenses_json)

    return {
        "insertRange": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": start_row - 1 + num_rows
            },
            "shiftDimension": "ROWS"
        }
    }




def get_mapped_cell_location(model_variable_mapping, section, field_key):
    for row in model_variable_mapping.to_dict(orient="records"):
        if row["section"].strip() == section.strip() and row["field_key"].strip() == field_key.strip():
            return row.get("location", "").replace("=", "")
    print(f"DEBUG: No location found for {section} {field_key}")
    return None

def get_operating_expenses_update_payload(
    expenses_json,
    rental_assumptions_json,
    amenity_income_json,
    model_variable_mapping,
    sheet_name="Operating Expenses",
    start_row=2
):

    expenses_json = sorted(
        expenses_json,
        key=lambda x: 0 if x["name"] == "Property Taxes" else (1 if x["name"] == "Insurance" else 2)
    )

    rental_row = len(rental_assumptions_json) + 2
    amenity_row = len(amenity_income_json) + 2
    rows_to_insert = []

    egi_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Estimated Pro Forma Rent Roll')
    # purchase_price_location = get_mapped_cell_location(model_variable_mapping, 'General Property Assumptions', 'Acquisition Price')
    # acquisition_loan_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Max Loan Size')
    gross_sf_location = get_mapped_cell_location(model_variable_mapping, 'General Property Assumptions', 'Gross Square Feet')
    net_rentable_sf_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Net Rentable SF')

    for i, expense in enumerate(expenses_json):
        row_num = start_row + i
        row = [""] * 10  # Columns A–J
        row[0] = expense["name"]       # Column A
        row[4] = expense["cost_per"]
        row[5] = expense["factor"]     # Expense column - factor is now cost per
        # row[9] = expense["broker"]     # Column J

        name = expense["name"]

        # Column G formula
        col_g_formula = ""

        if expense["cost_per"].lower() == "per unit":
            col_g_formula = f"='Rental Assumptions'!$D${rental_row}"

        #### FIX THESE
        elif expense["cost_per"].lower() == "per ca square foot":
            col_g_formula = f"={gross_sf_location}-{net_rentable_sf_location}"

        elif expense["cost_per"].lower() == "per total square feet":
            col_g_formula = f"={gross_sf_location}"

        elif expense["cost_per"].lower() == "percent of egi":
            col_g_formula = f'={egi_location}'
            try:
                row[5] = float(expense.get("factor") or 0) / 100
            except Exception:
                row[5] = 0

        # elif expense["cost_per"].lower() == "percent of purchase price":
        #     col_g_formula = f'={purchase_price_location}'
        #     row[5] = str(expense["factor"]) + "%"

        # elif expense["cost_per"].lower() == "percent of acquisition loan":
        #     col_g_formula = f'={acquisition_loan_location}'
        #     row[5] = str(expense["factor"]) + "%"

        row[6] = col_g_formula  # Column G

        # Column I (depends on G)
        if col_g_formula:
            row[8] = f"=F{row_num}*G{row_num}"
        else:
            row[8] = f"=F{row_num}"

        # Column H (depends on I)
        row[7] = f"=I{row_num}/12"

        rows_to_insert.append(row)

    end_row = start_row + len(expenses_json) - 1
    range_str = f"'{sheet_name}'!A{start_row}:J{end_row}"

    return {
        "range": range_str,
        "values": rows_to_insert
    }



def get_operating_expenses_format_payload(
    expenses_json,
    spreadsheet,
    sheet_name="Operating Expenses",
    start_row=2
):
    """
    Returns a list of format payloads for the operating expenses table.
    Applies number formatting for each row in the cost column (F) and statistic column (G).
    Ensures the correct sheet is referenced by name.
    """
    format_payloads = []
    sheet_id = get_sheet_id(spreadsheet, sheet_name)

    for i, expense in enumerate(expenses_json):
        row_num = start_row + i
        cost_per = str(expense.get("cost_per", "")).lower()

        # Column F (factor/cost value)
        if cost_per == "per unit":
            # No decimal places for per unit
            cost_number_format = {"type": "NUMBER", "pattern": "$#,##0\"/unit\""}
        elif cost_per in ["per ca square foot", "per total square feet"]:
            cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00\"/sf\""}
        elif cost_per == "per month":
            cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00\"/month\""}
        elif cost_per == "percent of egi":
            cost_number_format = {"type": "PERCENT", "pattern": "0.00%"}
        else:
            cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00"}

        format_payloads.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_num - 1,
                    "endRowIndex": row_num,
                    "startColumnIndex": 5,  # Column F (0-based)
                    "endColumnIndex": 6,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": cost_number_format,
                        "backgroundColor": {"red": 1, "green": 1, "blue": 1}
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.backgroundColor"
            }
        })

        # Column G (statistic)
        if cost_per == "per unit":
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" units\""}
        elif cost_per in ["per ca square foot", "per total square feet"]:
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" sf\""}
        elif cost_per == "per month":
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" months\""}
        else:
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0"}

        format_payloads.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_num - 1,
                    "endRowIndex": row_num,
                    "startColumnIndex": 6,  # Column G (0-based)
                    "endColumnIndex": 7,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": stat_number_format,
                        "backgroundColor": {"red": 1, "green": 1, "blue": 1}
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.backgroundColor"
            }
        })

        # Columns E and F text colored blue for inserted rows
        format_payloads.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_num - 1,
                    "endRowIndex": row_num,
                    "startColumnIndex": 4,  # Column E
                    "endColumnIndex": 6     # up to F (non-inclusive end)
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}},
                        "backgroundColor": {"red": 1, "green": 1, "blue": 1}
                    }
                },
                "fields": "userEnteredFormat.textFormat.foregroundColor,userEnteredFormat.backgroundColor"
            }
        })

    # Ensure all inserted rows are not bold (regular text) across a wide column range
    format_payloads.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": start_row - 1 + len(expenses_json),
                "startColumnIndex": 0,    # from column A
                "endColumnIndex": 200     # up to a wide bound (covers used columns)
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {
                        "bold": False
                    }
                }
            },
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })

    return format_payloads




def get_operating_expenses_sum_row_payload(
    expenses_json,
    sheet_name="Operating Expenses",
    start_row=2
):
    num_expenses = len(expenses_json)
    sum_row = start_row + num_expenses  # One row after the last inserted expense

    h_sum_formula = f"=SUM(H{start_row}:H{sum_row - 1})"
    i_sum_formula = f"=SUM(I{start_row}:I{sum_row - 1})"
    # j_sum_formula = f"=SUM(J{start_row}:J{sum_row - 1})"

    range_str = f"'{sheet_name}'!H{sum_row}:I{sum_row}"

    return {
        "range": range_str,
        "values": [[h_sum_formula, i_sum_formula]]
    }

def get_operating_expense_row_insert_request(
    spreadsheet,
    expenses_json,
    amenity_income_json,
    sheet_name="NOI Walk",
    start_base_row=34
):
    ws = spreadsheet.worksheet(sheet_name)
    start_row = start_base_row + len(amenity_income_json)
    num_expenses = len(expenses_json)

    return {
        "insertDimension": {
            "range": {
                "sheetId": ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_expenses
            },
            "inheritFromBefore": False
        }
    }



def get_operating_expense_row_insert_request_industrial(
    spreadsheet,
    expenses_json,
    amenity_income_json,
    sheet_name="NOI Walk",
    start_base_row=34
):
    ws = spreadsheet.worksheet(sheet_name)
    start_row = start_base_row + len(amenity_income_json)
    num_expenses = len(expenses_json)

    return {
        "insertDimension": {
            "range": {
                "sheetId": ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_expenses
            },
            "inheritFromBefore": False
        }
    }



def get_inflation_reference_updates(assumption_row_mapping):
    updates = []

    # Amenity Inflation → NOI Walk!L3
    amenity = assumption_row_mapping.get("Amenity Inflation")
    if amenity:
        formula = f"='Assumptions'!{amenity['value_col']}{amenity['row']}"
        updates.append({
            "range": "NOI Walk!L3",
            "values": [[formula]]
        })
    else:
        print("⚠️ Warning: Amenity Inflation not found in assumption_row_mapping")

    # Expense Inflation → NOI!H6
    expense = assumption_row_mapping.get("Expense Inflation")
    if expense:
        formula = f"='Assumptions'!{expense['value_col']}{expense['row']}"
        updates.append({
            "range": "NOI Walk!H6",
            "values": [[formula]]
        })
    else:
        print("⚠️ Warning: Expense Inflation not found in assumption_row_mapping")

    print("inflation_updates", updates)

    return updates

def get_operating_expense_formula_update_payloads(
    expenses_json,
    amenity_income_json,
    noi_sheet="NOI Walk",
    op_exp_sheet="Operating Expenses",
    start_base_row=34,
    start_col=5,
    num_months=132
):
    start_row = start_base_row + len(amenity_income_json)
    num_expenses = len(expenses_json)
    end_row = start_row + num_expenses - 1
    end_col_index = start_col + num_months - 1

    # Column B (labels from OpEx)
    b_values = [[f"='{op_exp_sheet}'!A{2 + i}"] for i in range(num_expenses)]
    b_range = f"'{noi_sheet}'!B{start_row}:B{end_row}"

    # Monthly formulas
    formula_rows = []
    egi_row = 27 + len(amenity_income_json)
    for i in range(num_expenses):
        op_row = 2 + i
        formula_row = []
        for j in range(num_months):
            col_index = start_col + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            # formula = f"='{op_exp_sheet}'!$H{op_row}*{col_letter}$11"

            formula = f"=IF('{op_exp_sheet}'!$E{op_row}=\"Percent of EGI\",'{op_exp_sheet}'!$F{op_row}*{col_letter}{egi_row},'{op_exp_sheet}'!$H{op_row}*{col_letter}$11)"

            formula_row.append(formula)
        formula_rows.append(formula_row)


    


    

   





    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, end_col_index).replace("1", "")
    e_range = f"'{noi_sheet}'!{start_col_letter}{start_row}:{end_col_letter}{end_row}"

    return [
        {"range": b_range, "values": b_values},
        {"range": e_range, "values": formula_rows}
    ]





def get_operating_expense_formula_update_payloads_industrial(
    expenses_json,
    amenity_income_json,
    noi_sheet="NOI Walk",
    op_exp_sheet="Retail Assumptions",
    start_base_row=34,
    start_col=5,
    num_months=132,
    ra_start_row=None
):
    start_row = start_base_row + len(amenity_income_json)
    num_expenses = len(expenses_json)
    end_row = start_row + num_expenses - 1
    end_col_index = start_col + num_months - 1

    # Column B (labels reference from Retail Assumptions expenses block, col B)
    if ra_start_row is None:
        ra_start_row = 17 + len(amenity_income_json) * 2  # fallback guess
    b_values = [[f"='{op_exp_sheet}'!B{ra_start_row + i}"] for i in range(num_expenses)]
    b_range = f"'{noi_sheet}'!B{start_row}:B{end_row}"

    # Monthly formulas
    formula_rows = []
    for i in range(num_expenses):
        formula_row = []
        for j in range(num_months):
            col_index = start_col + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            # Retail Assumptions monthlies start at column M (index 13)
            ra_col_letter = rowcol_to_a1(1, 14 + j).replace("1", "")
            ra_row = ra_start_row + i
            # Directly reference monthly totals from Retail Assumptions
            formula = f"='{op_exp_sheet}'!{ra_col_letter}{ra_row}"
            formula_row.append(formula)
        formula_rows.append(formula_row)


    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, end_col_index).replace("1", "")
    e_range = f"'{noi_sheet}'!{start_col_letter}{start_row}:{end_col_letter}{end_row}"

    return [
        {"range": b_range, "values": b_values},
        {"range": e_range, "values": formula_rows}
    ]

def get_expense_sum_row_to_noi_walk_payload(
    amenity_income_json,
    expenses_json,
    sheet_name="NOI Walk",
    start_col=5,
    num_months=132,
    row_offset=30
):
    start_sum_row = row_offset + len(amenity_income_json) + len(expenses_json)
    row_start = row_offset + len(amenity_income_json)
    row_end = row_start + len(expenses_json) - 1

    sum_row = []
    for j in range(num_months):
        col_index = start_col + j
        col_letter = rowcol_to_a1(1, col_index).replace("1", "")
        formula = f"=SUM({col_letter}{row_start}:{col_letter}{row_end})"
        sum_row.append(formula)

    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col + num_months - 1).replace("1", "")
    range_str = f"'{sheet_name}'!{start_col_letter}{start_sum_row}:{end_col_letter}{start_sum_row}"

    return {
        "range": range_str,
        "values": [sum_row]
    }



def get_expense_sum_row_to_noi_walk_payload_industrial(
    amenity_income_json,
    expenses_json,
    sheet_name="NOI Walk",
    start_col=5,
    num_months=132,
    start_base_row=23
):
    """
    Build the SUM row directly under the expense rows that were just inserted
    into NOI Walk via get_operating_expense_row_insert_request_industrial.
    Those inserts start at: start_row = start_base_row + len(amenity_income_json)
    and span len(expenses_json) rows.
    """
    # First expense row
    row_start = start_base_row + len(amenity_income_json)
    # Last expense row
    row_end = row_start + len(expenses_json) - 1
    # Sum row sits immediately after the last expense row
    start_sum_row = row_end + 1

    sum_row = []
    for j in range(num_months):
        col_index = start_col + j
        col_letter = rowcol_to_a1(1, col_index).replace("1", "")
        formula = f"=SUM({col_letter}{row_start}:{col_letter}{row_end})"
        sum_row.append(formula)

    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col + num_months - 1).replace("1", "")
    range_str = f"'{sheet_name}'!{start_col_letter}{start_sum_row}:{end_col_letter}{start_sum_row}"

    return {
        "range": range_str,
        "values": [sum_row]
    }

def get_noi_expense_rows_insert_and_update(
    spreadsheet,
    expenses_json,
    amenity_income_json,
    walk_start_row=30,
    noi_base_row=14,
    noi_sheet="NOI",
    walk_sheet="NOI Walk"
):
    num_rows = len(expenses_json)
    noi_start_row = noi_base_row + len(amenity_income_json)
    walk_start_row = walk_start_row + len(amenity_income_json)
    print("expenses_json", expenses_json)
    print("NUM ROWS", num_rows)
    # Insert request
    sheet_id = spreadsheet.worksheet(noi_sheet)._properties["sheetId"]
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": noi_start_row - 1,
                "endIndex": noi_start_row - 1 + num_rows
            },
            "inheritFromBefore": False
        }
    }

    reset_format = {
    "repeatCell": {
        "range": {
        "sheetId": sheet_id,
        "startRowIndex": noi_start_row - 1,
        "endRowIndex": noi_start_row - 1 + num_rows
        },
        "cell": {
        "userEnteredFormat": {
            "textFormat": { "bold": False },
            "backgroundColor": { "red": 1, "green": 1, "blue": 1 }
        }
        },
        "fields": "userEnteredFormat.textFormat.bold,userEnteredFormat.backgroundColor"
    }
    }

    # If there are no rows to insert, skip building rows and return empty updates safely
    if num_rows <= 0:
        try:
            print("[get_operating_expense_row_insert_request] num_rows == 0 → skipping values update")
        except Exception:
            pass
        # Return no insert, empty values payload, and no format reset
        return [], {"range": f"'{noi_sheet}'!A{noi_start_row}:A{noi_start_row - 1}", "values": []}, []

    # Values update (num_rows > 0 guaranteed below)
    all_rows = []
    for i in range(num_rows):
        walk_row = walk_start_row + i
        row = [""] * 2  # Columns A–B
        row.append(f"='{walk_sheet}'!B{walk_row}")  # Column C
        row += ["", "", ""]  # Columns D–F

        for col_letter in ["G", "H", "I", "J", "K"]:
            formula = (
                f"=SUMIFS('{walk_sheet}'!$E{walk_row}:$CW{walk_row},"
                f"'{walk_sheet}'!$E$14:$CW$14,"
                f"{noi_sheet}!{col_letter}$4)"
            )
            row.append(formula)

        all_rows.append(row)

    total_cols = len(all_rows[0])
    end_col_letter = rowcol_to_a1(1, total_cols).replace("1", "")
    end_row = noi_start_row + num_rows - 1
    range_str = f"'{noi_sheet}'!A{noi_start_row}:{end_col_letter}{end_row}"

    update_payload = {
        "range": range_str,
        "values": all_rows
    }

    return insert_request, update_payload, reset_format


def get_noi_expense_rows_insert_and_update_industrial(
    spreadsheet,
    expenses_json,
    amenity_income_json,
    walk_start_row=23,
    noi_base_row=11,
    noi_sheet="NOI",
    walk_sheet="NOI Walk"
):
    num_rows = len(expenses_json)
    noi_start_row = noi_base_row + len(amenity_income_json)
    walk_start_row = walk_start_row + len(amenity_income_json)
    # Insert request
    sheet_id = spreadsheet.worksheet(noi_sheet)._properties["sheetId"]
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": noi_start_row - 1,
                "endIndex": noi_start_row - 1 + num_rows
            },
            "inheritFromBefore": False
        }
    }

    reset_format = {
    "repeatCell": {
        "range": {
        "sheetId": sheet_id,
        "startRowIndex": noi_start_row - 1,
        "endRowIndex": noi_start_row - 1 + num_rows
        },
        "cell": {
        "userEnteredFormat": {
            "textFormat": { "bold": False },
            "backgroundColor": { "red": 1, "green": 1, "blue": 1 }
        }
        },
        "fields": "userEnteredFormat.textFormat.bold,userEnteredFormat.backgroundColor"
    }
    }

    # If there are no rows to insert, skip building rows and return empty updates safely
    if num_rows <= 0:
        try:
            print("[get_operating_expense_row_insert_request] num_rows == 0 → skipping values update")
        except Exception:
            pass
        # Return no insert, empty values payload, and no format reset
        return [], {"range": f"'{noi_sheet}'!A{noi_start_row}:A{noi_start_row - 1}", "values": []}, []

    # Values update (num_rows > 0 guaranteed below)
    all_rows = []
    for i in range(num_rows):
        walk_row = walk_start_row + i
        row = [""] * 2  # Columns A–B
        row.append(f"='{walk_sheet}'!B{walk_row}")  # Column C
        row += ["", "", ""]  # Columns D–F

        for col_letter in ["G", "H", "I", "J", "K"]:
            formula = (
                f"=SUMIFS('{walk_sheet}'!$E{walk_row}:$CW{walk_row},"
                f"'{walk_sheet}'!$E$13:$CW$13,"
                f"{noi_sheet}!{col_letter}$4)"
            )
            row.append(formula)

        all_rows.append(row)

    total_cols = len(all_rows[0])
    end_col_letter = rowcol_to_a1(1, total_cols).replace("1", "")
    end_row = noi_start_row + num_rows - 1
    range_str = f"'{noi_sheet}'!A{noi_start_row}:{end_col_letter}{end_row}"

    update_payload = {
        "range": range_str,
        "values": all_rows
    }

    return insert_request, update_payload, reset_format




def get_ntm_update_payload(
    amenity_income_json,
    expenses_json,
    sheet_name="NOI Walk",
    target_start_col=5,   # Column E
    target_num_cols=132,
    formula_start_col=6,  # Column F
    formula_range_width=11,  # F:Q = 11 columns
    row_offset=31
):
    row_1 = row_offset + len(amenity_income_json) + len(expenses_json) + 1
    target_row = row_offset + len(amenity_income_json) + len(expenses_json) + 2

    values = []
    for i in range(target_num_cols):
        start_idx = formula_start_col + i
        end_idx = start_idx + formula_range_width
        start_letter = rowcol_to_a1(1, start_idx).replace("1", "")
        end_letter = rowcol_to_a1(1, end_idx).replace("1", "")
        formula = f"=SUM({start_letter}{row_1}:{end_letter}{row_1})"
        values.append(formula)

    output_start_letter = rowcol_to_a1(1, target_start_col).replace("1", "")
    output_end_letter = rowcol_to_a1(1, target_start_col + target_num_cols - 1).replace("1", "")
    range_str = f"'{sheet_name}'!{output_start_letter}{target_row}:{output_end_letter}{target_row}"

    return {
        "range": range_str,
        "values": [values]
    }



def get_restabilization_update_payload(model_variable_mapping, rental_assumptions_json, sheet_name="Assumptions"):
    """
    Updates the Re-Stabilization Occurs cell with a dynamic formula based on rental units.
    Logic:
      - If no vacate flags = 1, return 1
      - Else take MAX lease-end where vacate=1, then add Lease-up Time + Rehab Time
    Uses SUMPRODUCT/LARGE to avoid Excel adding '@' on open.
    """
    # Target cell
    target_cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Re-Stabilization Occurs')
    if not target_cell:
        raise ValueError("Re-Stabilization Occurs location not found in model variable mapping")

    # Reference cells
    lease_up_time_cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Lease-up Time')
    rehab_time_cell    = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Rehab Time')
    if not lease_up_time_cell:
        raise ValueError("Lease-up Time location not found in model variable mapping")
    if not rehab_time_cell:
        raise ValueError("Rehab Time location not found in model variable mapping")

    # Dynamic ranges (adjust columns if yours differ)
    num_units = len(rental_assumptions_json)
    if num_units == 0:
        raise ValueError("No rental units found in rental_assumptions_json")
    end_row = 2 + num_units - 1

    # Vacate flags column (e.g., C) and Lease End column (e.g., F)
    vacate_range    = f"'Rental Assumptions'!C2:C{end_row}"
    lease_end_range = f"'Rental Assumptions'!F2:F{end_row}"


    # =IFERROR(IF(COUNTIF('Rental Assumptions'!D3:D4,1)=0,1,SUMPRODUCT(LARGE(('Rental Assumptions'!D3:D4=1)*'Rental Assumptions'!G3:G4,1))+Assumptions!G48+Assumptions!G47),0)

    # Build SUMPRODUCT/LARGE one-liner (no @ injection on open)
    formula = (
        f"=IFERROR("
        f"IF(COUNTIF({vacate_range},1)=0,1,"
        f"SUMPRODUCT(LARGE(({vacate_range}=1)*{lease_end_range},1))"
        f"+{lease_up_time_cell}+{rehab_time_cell}"
        f"),0)"
    )

    return {
        "range": f"{target_cell}",
        "values": [[formula]]
    }


def get_address_update_payload(address, model_variable_mapping, sheet_name="Assumptions"):
    cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Address')
    if not cell:
        raise ValueError("Address location not found in model variable mapping")
    
    range_str = f"'{sheet_name}'!{cell}" if "!" not in str(cell) else cell
    
    return {
        "range": range_str,
        "values": [[address]]
    }

def get_property_name_update_payload(property_name, model_variable_mapping, sheet_name="Cover"):
    cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Property Name')
    if not cell:
        range_str = f"'{sheet_name}'!A1"
    else:
        range_str = f"'{sheet_name}'!{cell}" if "!" not in str(cell) else cell
        
    return {
        "range": range_str,
        "values": [[property_name]]
    }


def get_number_of_spaces_update_payload(model_variable_mapping, retail_income_json, sheet_name="Cover"):
    cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Number of Spaces')

    insert_function = f"=COUNTA(UNIQUE('{sheet_name}'!$B$6:B{len(retail_income_json) + 5}))"
    if not cell:
        return []
        
    range_str = f"{cell}"
    return {
        "range": range_str,
        "values": [[insert_function]]
    }

def get_in_place_rent_update_payload(model_variable_mapping, retail_income_json, sheet_name="Cover"):
    cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'In-Place Rent / SF')

    insert_function = (
        f'=+SUMIFS('
        f'\'Retail Assumptions\'!$L$6:$L${len(retail_income_json) + 5},'
        f'\'Retail Assumptions\'!$G$6:$G${len(retail_income_json) + 5},">="&1)'
        f'/'
        f'SUMIFS('
        f'\'Retail Assumptions\'!$F$6:$F${len(retail_income_json) + 5},'
        f'\'Retail Assumptions\'!$G$6:$G${len(retail_income_json) + 5},">="&1)'
    )
    if not cell:
        return []
        
    range_str = f"{cell}"
    return {
        "range": range_str,
        "values": [[insert_function]]
    }

def get_retail_assumptions_inserts(spreadsheet, retail_income, sheet_name="Retail Assumptions", start_row=6):
    """
    Returns insert and update requests for retail income data in the Retail Assumptions sheet.
    
    Args:
        spreadsheet: The gspread spreadsheet object
        retail_income: List of retail income dictionaries
        sheet_name: Name of the sheet to update (default: "Retail Assumptions")
        start_row: Starting row for insertions (default: 6)
    
    Returns:
        tuple: (insert_request, update_payload, format_requests)
    """
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    num_rows = len(retail_income)
    end_row = start_row + num_rows - 1
    
    # Insert request - insert rows starting at start_row
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + num_rows
            },
            "inheritFromBefore": False
        }
    }
    
    # Build rows for update - need to extend range to include all columns we're writing to
    rows = []
    for i, entry in enumerate(retail_income):
        current_row = start_row + i
        row = [""] * 145  # Columns A through EQ (145 columns total)
        
        # Column B: suite
        row[1] = entry.get("suite", "")
        
        # Column C: tenant_name  
        row[2] = entry.get("tenant_name", "")
        
        # Column D: blank
        row[3] = ""
        
        # Column E: lease_start_month
        row[4] = entry.get("lease_start_month", "")
        
        # Column F: square_feet
        row[5] = entry.get("square_feet", "")
        
        # Column G: lease_start_month (same as E)
        row[6] = entry.get("lease_start_month", "")
        
        # Column H: annual_bumps (with percent sign)
        annual_bumps = entry.get("annual_bumps", "")
        if annual_bumps != "":
            row[7] = f"{annual_bumps}%"
        else:
            row[7] = ""
        
        # Column I: rent_per_square_foot_per_year
        row[8] = entry.get("rent_per_square_foot_per_year", "")
        
        # Column J: Formula I<row> * F<row>
        row[9] = f"=I{current_row}*F{current_row}"
        
        # Columns L through EQ (columns 11 through 144): Formula pattern
        for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=($G{current_row}<={col_letter}$5)*($J{current_row}/12)*(1+$H{current_row})^(ROUNDUP(MAX({col_letter}$5-$G{current_row}+1,0)/12,0)-1)"
        
        rows.append(row)
    
    # Update payload - extend range to cover all columns we're writing to (A to EQ)
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{end_row}",
        "values": rows
    }

    # Format requests: make B–J non-bold, and color B–I blue for inserted rows
    # Note: Google Sheets uses 0-based indices. B=1, I=8, J=9 (endColumnIndex is exclusive)
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 9    # up to I (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    return insert_request, update_payload, format_requests



def get_retail_assumptions_summary_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail assumptions data.
    This row goes after all the retail income entries.
    """
    start_row = 6 + len(retail_income)  # Row after all retail entries
    end_row = start_row
    
    # Create the summary row with formulas
    summary_row = [""] * 145  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Base Retail Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 6 + len(retail_income) - 1
    summary_row[4] = f"=MIN(E6:E{data_end_row})"
    
    # Column F: =SUM(F6:F<6+len(retail_income))
    summary_row[5] = f"=SUM(F6:F{data_end_row})"
    
    # Column G: =MIN(G6:G<6+len(retail_income))
    summary_row[6] = f"=MIN(G6:G{data_end_row})"
    
    # Column H: =IFERROR(SUMPRODUCT(H6:H<data_end>,J6:J<data_end>)/J<summary_row>,0)
    summary_row[7] = f"=IFERROR(SUMPRODUCT(H6:H{data_end_row},J6:J{data_end_row})/J{start_row},0)"
    
    # Column I: =IFERROR(SUMPRODUCT(I6:I<data_end>,F6:F<data_end>)/F<summary_row>,0)
    summary_row[8] = f"=IFERROR(SUMPRODUCT(I6:I{data_end_row},F6:F{data_end_row})/F{start_row},0)"
    
    # Column J: =SUM(J6:J<data_end>)
    summary_row[9] = f"=SUM(J6:J{data_end_row})"
    
    # Columns L through EQ (columns 11 through 144): Sum formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}6:{col_letter}{data_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload



def get_retail_recovery_summary_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail recovery data.
    This row goes after all the retail recovery entries.
    """
    start_row = 10 + len(retail_income)*2  # Row after all retail entries
    end_row = start_row
    data_start_row = 10 + len(retail_income)
    
    # Create the summary row with formulas
    summary_row = [""] * 145  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Recovery Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 10 + len(retail_income)*2 
    
    # Column I: 
    summary_row[8] = f"=IFERROR(J{data_end_row}/F{6 + len(retail_income)},0)"
    
    # Column J: =SUM(J6:J<data_end>)
    summary_row[9] = f"=SUM(J{10 + len(retail_income)}:J{data_end_row -1})"
    
    # Columns L through EQ (columns 11 through 144): Sum formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row -1})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload

def get_retail_assumptions_occ_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates an occ row below the summary row with SUMIF formulas.
    """
    start_row = 6 + len(retail_income) + 1  # Row after summary row
    data_end_row = 6 + len(retail_income) - 1  # Last data row
    summary_row_num = 6 + len(retail_income)  # Summary row number
    
    # Create the occ row with formulas
    occ_row = [""] * 145  # Initialize with empty strings for all columns
    
    occ_row[10] = "Occ."
    
    # Columns L through EO (columns 11 through 144): SUMIF formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        occ_row[col_idx] = f"=IFERROR(SUMIF($E$6:$E${data_end_row},\"<=\"&{col_letter}5,$F$6:$F${data_end_row})/$F${summary_row_num},0)"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [occ_row]
    }
    
    return update_payload


def get_retail_recovery_inserts(spreadsheet, retail_income, retail_expenses, sheet_name='Retail Assumptions'):
    """
    Creates insert request and update payload for retail recovery rows.
    """
    retail_income_length = len(retail_income)
    start_row = 10 + retail_income_length  # Starting row for recovery section
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    # Insert request for new rows
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_income_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    for i, item in enumerate(retail_income):
        row = [""] * 145  # Initialize with empty strings for all columns
        
        # Column B: reference to previous table (=+B6, =+B7, etc.)
        row[1] = f"=+B{6 + i}"
        
        # Column C: reference to previous table (=+C6, =+C7, etc.)
        row[2] = f"=+C{6 + i}"
        
        # Column F: recovery_start_month
        row[5] = item.get("recovery_start_month", "")
        
        # Column H: =IFERROR(F{current_row}/$F${summary_row},0)
        current_row = 10 + i + retail_income_length
        summary_row = 6 + retail_income_length  # Summary row from previous table
        upper_row = 6 + i 
        row[7] = f"=IFERROR(F{upper_row}/$F${summary_row},0)"
        print(current_row)

        # Column I: =IFERROR(J{current_row}/F{prev_table_row},0)
        prev_table_row = 6 + i
        row[8] = f"=IFERROR(J{current_row}/F{prev_table_row},0)"
        
        # Column J: =+H{current_row}*$J${reference_row}
        reference_row = 17 + retail_income_length * 2 + len(retail_expenses)
        row[9] = f"=+H{current_row}*$J${reference_row}"
        
        # Columns L through EO (columns 11 through 144): =IFERROR((COL$5>=$F{current_row})*$H{current_row}*COL${reference_row},0)
        for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=IFERROR(({col_letter}$5>=$F{current_row})*$H{current_row}*{col_letter}${reference_row},0)"
        
        data_rows.append(row)
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row + retail_income_length - 1}",
        "values": data_rows
    }

    # Formatting: all inserted row values B–J non-bold; only B, C, F blue
    end_row = start_row + retail_income_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 3    # C (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 5,  # F
                    "endColumnIndex": 6
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            # Format Column I (Pro Rata Share) as percentage
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "PERCENT",
                            "pattern": "0.00%"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        }
    ]

    return insert_request, update_payload, format_requests


def get_retail_expenses_inserts(spreadsheet, retail_income, retail_expenses, retail_growth_rate, sheet_name):
    """
    Insert retail expenses rows and update with data
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate starting row
    start_row = 17 + retail_income_length * 2

    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    # Create insert request
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_expenses_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    for i, expense in enumerate(retail_expenses):
        row = [""] * 145  # Initialize with empty strings for all columns
        current_row = start_row + i
        
        # Column B: expense name
        row[1] = expense.get("name", "")
        
        # Column G: retail growth rate
        row[6] = f"{retail_growth_rate}%"
        
        # Column I: cost_per
        row[8] = expense.get("cost_per", "")
        
        # Column J: =F{6 + len(retail_expenses)} * I{current_row}
        reference_row = 6 + retail_income_length
        row[9] = f"=F{reference_row}*I{current_row}"
        
        # Columns L through EO (columns 11 through 144): =IFERROR($J{current_row}/12*COL${reference_row_20}*(1+$G{current_row})^(ROUNDUP(COL$5/12,0)-1),0)
        reference_occ_row = 6 + retail_income_length + 1  
        for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=IFERROR($J{current_row}/12*{col_letter}${reference_occ_row}*(1+$G{current_row})^(ROUNDUP({col_letter}$5/12,0)-1),0)"
        
        data_rows.append(row)
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EO{start_row + retail_expenses_length - 1}",
        "values": data_rows
    }

    # Formatting for inserted rows: B–J non-bold; only column I blue
    end_row = start_row + retail_expenses_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    return insert_request, update_payload, format_requests

def get_retail_expenses_summary_row(retail_income, retail_expenses, sheet_name):
    """
    Creates the summary row for retail expenses.
    Row position: 17 + len(retail_income)*2 + len(retail_expenses)
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate row position
    summary_row = 17 + retail_income_length * 2 + retail_expenses_length
    
    # Initialize row with empty strings
    row = [""] * 145
    
    # Column B: text
    row[1] = "Total Retail Operating Expenses"
    
    # Column I: =IFERROR(J{summary_row}/F{6 + retail_income_length},0)
    reference_row = 6 + retail_income_length
    row[8] = f"=IFERROR(J{summary_row}/F{reference_row},0)"
    
    # Column J: SUM of retail expenses column J
    expenses_start_row = 17 + retail_income_length * 2
    expenses_end_row = expenses_start_row + retail_expenses_length - 1
    row[9] = f"=SUM(J{expenses_start_row}:J{expenses_end_row})"
    
    # Columns L through EO (columns 11 through 144): SUM of expense rows for that column
    for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        row[col_idx] = f"=SUM({col_letter}{expenses_start_row}:{col_letter}{expenses_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{summary_row}:EO{summary_row}",
        "values": [row]
    }
    
    return update_payload




def get_retail_assumptions_inserts(spreadsheet, retail_income, sheet_name="Retail Assumptions", start_row=6):
    """
    Returns insert and update requests for retail income data in the Retail Assumptions sheet.
    
    Args:
        spreadsheet: The gspread spreadsheet object
        retail_income: List of retail income dictionaries
        sheet_name: Name of the sheet to update (default: "Retail Assumptions")
        start_row: Starting row for insertions (default: 6)
    
    Returns:
        tuple: (insert_request, update_payload, format_requests)
    """
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    num_rows = len(retail_income)
    end_row = start_row + num_rows - 1
    
    # Insert request - insert rows starting at start_row
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + num_rows
            },
            "inheritFromBefore": False
        }
    }
    
    # Build rows for update - need to extend range to include all columns we're writing to
    rows = []
    for i, entry in enumerate(retail_income):
        current_row = start_row + i
        row = [""] * 145  # Columns A through EQ (145 columns total)
        
        # Column B: suite
        row[1] = entry.get("suite", "")
        
        # Column C: tenant_name  
        row[2] = entry.get("tenant_name", "")
        
        # Column D: blank
        row[3] = ""
        
        # Column E: lease_start_month
        row[4] = entry.get("lease_start_month", "")
        
        # Column F: square_feet
        row[5] = entry.get("square_feet", "")
        
        # Column G: lease_start_month (same as E)
        row[6] = entry.get("lease_start_month", "")
        
        # Column H: annual_bumps (with percent sign)
        annual_bumps = entry.get("annual_bumps", "")
        if annual_bumps != "":
            row[7] = f"{annual_bumps}%"
        else:
            row[7] = ""
        
        # Column I: rent_per_square_foot_per_year
        row[8] = entry.get("rent_per_square_foot_per_year", "")
        
        # Column J: Formula I<row> * F<row>
        row[9] = f"=I{current_row}*F{current_row}"
        
        # Columns L through EQ (columns 11 through 144): Formula pattern
        for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=($G{current_row}<={col_letter}$5)*($J{current_row}/12)*(1+$H{current_row})^(ROUNDUP(MAX({col_letter}$5-$G{current_row}+1,0)/12,0)-1)"
        
        rows.append(row)
    
    # Update payload - extend range to cover all columns we're writing to (A to EQ)
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{end_row}",
        "values": rows
    }

    # Format requests: make B–J non-bold, and color B–I blue for inserted rows
    # Note: Google Sheets uses 0-based indices. B=1, I=8, J=9 (endColumnIndex is exclusive)
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 9    # up to I (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    return insert_request, update_payload, format_requests



def get_retail_assumptions_summary_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail assumptions data.
    This row goes after all the retail income entries.
    """
    start_row = 6 + len(retail_income)  # Row after all retail entries
    end_row = start_row
    
    # Create the summary row with formulas
    summary_row = [""] * 145  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Base Retail Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 6 + len(retail_income) - 1
    summary_row[4] = f"=MIN(E6:E{data_end_row})"
    
    # Column F: =SUM(F6:F<6+len(retail_income))
    summary_row[5] = f"=SUM(F6:F{data_end_row})"
    
    # Column G: =MIN(G6:G<6+len(retail_income))
    summary_row[6] = f"=MIN(G6:G{data_end_row})"
    
    # Column H: =IFERROR(SUMPRODUCT(H6:H<data_end>,J6:J<data_end>)/J<summary_row>,0)
    summary_row[7] = f"=IFERROR(SUMPRODUCT(H6:H{data_end_row},J6:J{data_end_row})/J{start_row},0)"
    
    # Column I: =IFERROR(SUMPRODUCT(I6:I<data_end>,F6:F<data_end>)/F<summary_row>,0)
    summary_row[8] = f"=IFERROR(SUMPRODUCT(I6:I{data_end_row},F6:F{data_end_row})/F{start_row},0)"
    
    # Column J: =SUM(J6:J<data_end>)
    summary_row[9] = f"=SUM(J6:J{data_end_row})"
    
    # Columns L through EQ (columns 11 through 144): Sum formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}6:{col_letter}{data_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload



def get_retail_recovery_summary_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail recovery data.
    This row goes after all the retail recovery entries.
    """
    start_row = 10 + len(retail_income)*2  # Row after all retail entries
    end_row = start_row
    data_start_row = 10 + len(retail_income)
    
    # Create the summary row with formulas
    summary_row = [""] * 145  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Recovery Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 10 + len(retail_income)*2 
    
    # Column I: 
    summary_row[8] = f"=IFERROR(J{data_end_row}/F{6 + len(retail_income)},0)"
    
    # Column J: =SUM(J6:J<data_end>)
    summary_row[9] = f"=SUM(J{10 + len(retail_income)}:J{data_end_row -1})"
    
    # Columns L through EQ (columns 11 through 144): Sum formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row -1})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload

def get_retail_assumptions_occ_row(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates an occ row below the summary row with SUMIF formulas.
    """
    start_row = 6 + len(retail_income) + 1  # Row after summary row
    data_end_row = 6 + len(retail_income) - 1  # Last data row
    summary_row_num = 6 + len(retail_income)  # Summary row number
    
    # Create the occ row with formulas
    occ_row = [""] * 145  # Initialize with empty strings for all columns
    
    occ_row[10] = "Occ."
    
    # Columns L through EO (columns 11 through 144): SUMIF formulas
    for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        occ_row[col_idx] = f"=IFERROR(SUMIF($E$6:$E${data_end_row},\"<=\"&{col_letter}5,$F$6:$F${data_end_row})/$F${summary_row_num},0)"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [occ_row]
    }
    
    return update_payload


def get_retail_recovery_inserts(spreadsheet, retail_income, retail_expenses, sheet_name='Retail Assumptions'):
    """
    Creates insert request and update payload for retail recovery rows.
    """
    retail_income_length = len(retail_income)
    start_row = 10 + retail_income_length  # Starting row for recovery section
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    # Insert request for new rows
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_income_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    for i, item in enumerate(retail_income):
        row = [""] * 145  # Initialize with empty strings for all columns
        
        # Column B: reference to previous table (=+B6, =+B7, etc.)
        row[1] = f"=+B{6 + i}"
        
        # Column C: reference to previous table (=+C6, =+C7, etc.)
        row[2] = f"=+C{6 + i}"
        
        # Column F: recovery_start_month
        row[5] = item.get("recovery_start_month", "")
        
        # Column H: =IFERROR(F{current_row}/$F${summary_row},0)
        current_row = 10 + i + retail_income_length
        summary_row = 6 + retail_income_length  # Summary row from previous table
        upper_row = 6 + i 
        row[7] = f"=IFERROR(F{upper_row}/$F${summary_row},0)"
        print(current_row)

        # Column I: =IFERROR(J{current_row}/F{prev_table_row},0)
        prev_table_row = 6 + i
        row[8] = f"=IFERROR(J{current_row}/F{prev_table_row},0)"
        
        # Column J: =+H{current_row}*$J${reference_row}
        reference_row = 17 + retail_income_length * 2 + len(retail_expenses)
        row[9] = f"=+H{current_row}*$J${reference_row}"
        
        # Columns L through EO (columns 11 through 144): =IFERROR((COL$5>=$F{current_row})*$H{current_row}*COL${reference_row},0)
        for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=IFERROR(({col_letter}$5>=$F{current_row})*$H{current_row}*{col_letter}${reference_row},0)"
        
        data_rows.append(row)
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row + retail_income_length - 1}",
        "values": data_rows
    }

    # Formatting: all inserted row values B–J non-bold; only B, C, F blue
    end_row = start_row + retail_income_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 3    # C (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 5,  # F
                    "endColumnIndex": 6
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    return insert_request, update_payload, format_requests


def get_retail_expenses_inserts(spreadsheet, retail_income, retail_expenses, retail_growth_rate, sheet_name):
    """
    Insert retail expenses rows and update with data
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate starting row
    start_row = 17 + retail_income_length * 2

    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    # Create insert request
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_expenses_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    for i, expense in enumerate(retail_expenses):
        row = [""] * 145  # Initialize with empty strings for all columns
        current_row = start_row + i
        
        # Column B: expense name
        row[1] = expense.get("name", "")
        
        # Column G: retail growth rate
        row[6] = f"{retail_growth_rate}%"
        
        # Column I: cost_per
        row[8] = expense.get("cost_per", "")
        
        # Column J: =F{6 + len(retail_expenses)} * I{current_row}
        reference_row = 6 + retail_income_length
        row[9] = f"=F{reference_row}*I{current_row}"
        
        # Columns L through EO (columns 11 through 144): =IFERROR($J{current_row}/12*COL${reference_row_20}*(1+$G{current_row})^(ROUNDUP(COL$5/12,0)-1),0)
        reference_occ_row = 6 + retail_income_length + 1  
        for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=IFERROR($J{current_row}/12*{col_letter}${reference_occ_row}*(1+$G{current_row})^(ROUNDUP({col_letter}$5/12,0)-1),0)"
        
        data_rows.append(row)
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EO{start_row + retail_expenses_length - 1}",
        "values": data_rows
    }

    # Formatting for inserted rows: B–J non-bold; only column I blue
    end_row = start_row + retail_expenses_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 10   # up to J (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    return insert_request, update_payload, format_requests

def get_retail_expenses_summary_row(retail_income, retail_expenses, sheet_name):
    """
    Creates the summary row for retail expenses.
    Row position: 17 + len(retail_income)*2 + len(retail_expenses)
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate row position
    summary_row = 17 + retail_income_length * 2 + retail_expenses_length
    
    # Initialize row with empty strings
    row = [""] * 145
    
    # Column B: text
    row[1] = "Total Retail Operating Expenses"
    
    # Column I: =IFERROR(J{summary_row}/F{6 + retail_income_length},0)
    reference_row = 6 + retail_income_length
    row[8] = f"=IFERROR(J{summary_row}/F{reference_row},0)"
    
    # Column J: SUM of retail expenses column J
    expenses_start_row = 17 + retail_income_length * 2
    expenses_end_row = expenses_start_row + retail_expenses_length - 1
    row[9] = f"=SUM(J{expenses_start_row}:J{expenses_end_row})"
    
    # Columns L through EO (columns 11 through 144): SUM of expense rows for that column
    for col_idx in range(11, 145):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        row[col_idx] = f"=SUM({col_letter}{expenses_start_row}:{col_letter}{expenses_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{summary_row}:EO{summary_row}",
        "values": [row]
    }
    
    return update_payload

    
def get_retail_assumptions_inserts_industrial(spreadsheet, retail_income, sheet_name="Retail Assumptions", start_row=6):
    """
    Returns insert and update requests for retail income data in the Retail Assumptions sheet.
    
    Args:
        spreadsheet: The gspread spreadsheet object
        retail_income: List of retail income dictionaries
        sheet_name: Name of the sheet to update (default: "Retail Assumptions")
        start_row: Starting row for insertions (default: 6)
    
    Returns:
        tuple: (insert_request, update_payload, format_requests)
    """
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    num_rows = len(retail_income)
    end_row = start_row + num_rows - 1
    
    # Insert request - insert rows starting at start_row
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + num_rows
            },
            "inheritFromBefore": False
        }
    }
    
    # Build rows for update - need to extend range to include all columns we're writing to
    rows = []
    for i, entry in enumerate(retail_income):
        current_row = start_row + i
        row = [""] * 147 # Columns A through EQ (145 columns total)
        
        # Column B: suite
        row[1] = entry.get("suite", "")
        
        # Column C: tenant_name  
        row[2] = entry.get("tenant_name", "")
        
        # Column D: blank
        row[3] = ""
        
        # Column E: lease_start_month
        row[4] = entry.get("lease_start_month", "")
        
        # Column F: square_feet
        row[5] = entry.get("square_feet", "")
        
        # Column G: lease_start_month (same as E)
        row[6] = entry.get("lease_start_month", "")

        # Column H: lease_end_month (new)
        row[7] = entry.get("lease_end_month", "")

        # Column I: annual_bumps (with percent sign) (shifted)
        annual_bumps = entry.get("annual_bumps", "")
        row[8] = f"{annual_bumps}%" if annual_bumps != "" else ""
        
        # Column J: rent_per_square_foot_per_year (shifted)
        row[9] = entry.get("rent_per_square_foot_per_year", "")
        
        # Column K: rent_type (new)
        row[10] = entry.get("rent_type", "")

        # Column L: Annual Rent = J * F (shifted)
        row[11] = f"=J{current_row}*F{current_row}"
        
        # Columns M through EQ (columns 12 through 144): Formula pattern
        for col_idx in range(13, 147):  # M is column 12 (0-based), EQ is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            # Use L (Annual Rent) and I (Annual Bumps) after column additions
            row[col_idx] = f"=($G{current_row}<={col_letter}$5)*($L{current_row}/12)*(1+$I{current_row})^(ROUNDUP(MAX({col_letter}$5-$G{current_row}+1,0)/12,0)-1)"
        
        rows.append(row)
    
    # Update payload - extend range to cover all columns we're writing to (A to EQ)
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{end_row}",
        "values": rows
    }

    # Format requests: make B–L non-bold, and color B–K blue for inserted rows
    # Note: Google Sheets uses 0-based indices. B=1, K=10, L=11 (endColumnIndex is exclusive)
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,   # B
                    "endColumnIndex": 12     # up to L (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,   # B
                    "endColumnIndex": 11     # up to K (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]

    # Ensure all inserted rows (full width A..EQ) are regular (not bold)
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": end_row,
                "startColumnIndex": 0,   # A
                "endColumnIndex": 147    # EQ is 147th column (0-based end exclusive)
            },
            "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })

    return insert_request, update_payload, format_requests



def get_retail_assumptions_summary_row_industrial(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail assumptions data.
    This row goes after all the retail income entries.
    """
    start_row = 6 + len(retail_income)  # Row after all retail entries
    end_row = start_row
    
    # Create the summary row with formulas
    summary_row = [""] * 147  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Base Retail Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 6 + len(retail_income) - 1
    summary_row[4] = f"=MIN(E6:E{data_end_row})"
    
    # Column F: =SUM(F6:F<6+len(retail_income))
    summary_row[5] = f"=SUM(F6:F{data_end_row})"
    
    # Column G: =MIN(G6:G<6+len(retail_income))
    summary_row[6] = f"=MAX(G6:G{data_end_row})"
    
    # Column H (Lease End): =MAX(H6:H<data_end>)
    summary_row[7] = f"=MAX(H6:H{data_end_row})"
    
    # Column I (Annual Bumps): weighted average by SF
    summary_row[8] = f"=IFERROR(SUMPRODUCT(I6:I{data_end_row},F6:F{data_end_row})/F{start_row},0)"
    
    # Column J (Rent / SF / Yr.): weighted average by SF
    summary_row[9] = f"=IFERROR(SUMPRODUCT(J6:J{data_end_row},F6:F{data_end_row})/F{start_row},0)"
    
    # Column L (Annual Rent): sum of annual rent
    summary_row[11] = f"=SUM(L6:L{data_end_row})"
    
    # Columns M through EQ (columns 12 through 144): Sum formulas for monthly values
    for col_idx in range(13, 147):  # M is column 12 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}6:{col_letter}{data_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload



def get_retail_recovery_summary_row_industrial(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates a summary row with formulas for the retail recovery data.
    This row goes after all the retail recovery entries.
    """
    start_row = 10 + len(retail_income)*2  # Row after all retail entries
    end_row = start_row
    data_start_row = 10 + len(retail_income)
    
    # Create the summary row with formulas
    summary_row = [""] * 147  # Initialize with empty strings for all columns
    
    summary_row[1] = "Total Recovery Income"
    
    # Column E: =MIN(E6:E<6+len(retail_income))
    data_end_row = 10 + len(retail_income)*2 
    
    # # Column I: 
    # summary_row[8] = f"=IFERROR(J{data_end_row}/F{6 + len(retail_income)},0)"
    
    # Column J: =SUM(J6:J<data_end>)
    summary_row[9] = f"=SUM(J{10 + len(retail_income)}:J{data_end_row -1})"
    
    # Columns L through EQ (columns 11 through 144): Sum formulas
    for col_idx in range(13, 147):  # L is column 11 (0-based), EQ is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        summary_row[col_idx] = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row -1})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [summary_row]
    }
    
    return update_payload

def get_retail_assumptions_occ_row_industrial(retail_income, sheet_name='Retail Assumptions'):
    """
    Creates an occ row below the summary row with SUMIF formulas.
    """
    start_row = 6 + len(retail_income) + 1  # Row after summary row
    data_end_row = 6 + len(retail_income) - 1  # Last data row
    summary_row_num = 6 + len(retail_income)  # Summary row number
    
    # Create the occ row with formulas
    occ_row = [""] * 147  # Initialize with empty strings for all columns
    
    occ_row[12] = "Occ."
    
    # Columns L through EO (columns 11 through 144): SUMIF formulas
    for col_idx in range(13, 147):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        occ_row[col_idx] = f"=IFERROR(SUMIF($E$6:$E${data_end_row},\"<=\"&{col_letter}5,$F$6:$F${data_end_row})/$F${summary_row_num},0)"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row}",
        "values": [occ_row]
    }
    
    return update_payload


def get_retail_recovery_inserts_industrial(spreadsheet, retail_income, retail_expenses, sheet_name='Retail Assumptions'):
    """
    Creates insert request and update payload for retail recovery rows.
    """
    retail_income_length = len(retail_income)
    start_row = 10 + retail_income_length  # Starting row for recovery section
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    # Insert request for new rows
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_income_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    for i, item in enumerate(retail_income):
        row = [""] * 147  # Initialize with empty strings for all columns
        
        # Column B: reference to previous table (=+B6, =+B7, etc.)
        row[1] = f"=+B{6 + i}"
        
        # Column C: reference to previous table (=+C6, =+C7, etc.)
        row[2] = f"=+C{6 + i}"
        
        # Column E: Lease Start (reference from base table)
        row[4] = f"=+E{6 + i}"
        
        # Column F: recovery_start_month
        row[5] = item.get("recovery_start_month", "")
        
        current_row = 10 + i + retail_income_length
        summary_row = 6 + retail_income_length  # Summary row from previous table
        upper_row = 6 + i 

        # Column H: Lease End (reference from base table)
        row[7] = f"=+H{upper_row}"
        
        # Column I: Pro Rata Share =IFERROR(F{upper_row}/$F${summary_row},0)
        current_row = 10 + i + retail_income_length
        row[8] = f"=IFERROR(F{upper_row}/$F${summary_row},0)"

        # Column J: Rec. / SF / Yr. = Annual (L) / SF (from base table col F)
        # Example: =IFERROR(L26/F9,0)
        prev_table_row = 6 + i
        reference_row = 17 + retail_income_length * 2 + len(retail_expenses)
        row[9] = f"=IFERROR(L{current_row}/F{upper_row},0)"
        
        # Column K: Rent Type (reference from base table)
        row[10] = f"=+K{upper_row}"
        
        # Column L: Annual Recovery
        # = (SUMIFS(L$start:L$end, $K$start:$K$end, K{current_row}) + SUMIFS(L$start:L$end, $K$start:$K$end, "Both")) * $I{current_row}
        expenses_start_row = 17 + retail_income_length * 2
        expenses_end_row = expenses_start_row + len(retail_expenses) - 1
        row[11] = (
            f"=(SUMIFS(L${expenses_start_row}:L${expenses_end_row},$K${expenses_start_row}:$K${expenses_end_row},K{current_row})"
            f"+SUMIFS(L${expenses_start_row}:L${expenses_end_row},$K${expenses_start_row}:$K${expenses_end_row},\"Both\"))*$I{current_row}"
        )
        
        # Columns M through EO (columns 12 through 144):
        # =IFERROR((COL$5>=$F{current_row})*$I{current_row}*COL${reference_row},0)
        for col_idx in range(13, 147):  # M is column 12 (0-based), EO is column 144
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            row[col_idx] = f"=IFERROR(({col_letter}$5>=$F{current_row})*$I{current_row}*{col_letter}${reference_row},0)"
        
        data_rows.append(row)
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row + retail_income_length - 1}",
        "values": data_rows
    }

    # Formatting: all inserted row values B–L non-bold; only B, C, F blue
    end_row = start_row + retail_income_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,   # B
                    "endColumnIndex": 12     # up to L (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 3    # C (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 5,  # F
                    "endColumnIndex": 6
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        },
        {
            # Format Column I (Pro Rata Share) as percentage
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "PERCENT",
                            "pattern": "0.00%"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        }
    ]

    return insert_request, update_payload, format_requests


def get_retail_expenses_inserts_industrial(spreadsheet, retail_income, retail_expenses, retail_growth_rate, sheet_name, model_variable_mapping=None):
    """
    Insert retail expenses rows and update with data
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate starting row
    start_row = 17 + retail_income_length * 2

    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    
    # Create insert request
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": start_row - 1,  # 0-based index
                "endIndex": start_row - 1 + retail_expenses_length
            }
        }
    }
    
    # Create data rows
    data_rows = []
    per_row_formats: list = []
    for i, expense in enumerate(retail_expenses):
        row = [""] * 147  # Initialize with empty strings for all columns
        current_row = start_row + i
        
        # Column B: expense name
        row[1] = expense.get("name", "")

        # Column E: Expense Growth -> mapped cell (fallback to provided percent)
        if model_variable_mapping is not None:
            print("EXPENSE INFLATION")
            try:
                growth_cell = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Expense Inflation')
                print("GROWTH CELL", growth_cell)
                row[4] = f"={growth_cell}"
            except Exception:
                row[4] = f"{retail_growth_rate}%"
        else:
            row[4] = f"{retail_growth_rate}%"

        # Column G: Factor (text)
        factor_text = str(expense.get("factor", "Annual"))
        row[6] = factor_text
        
        # Column I: Expense (cost_per) - ensure percent factors are converted to fraction
        raw_cost = expense.get("cost_per", "")
        normalized_cost = raw_cost
        try:
            import re
            if isinstance(raw_cost, str):
                # strip out symbols like $,%, and commas
                num_str = re.sub(r"[^\d.\-]", "", raw_cost)
                normalized_cost = float(num_str) if num_str != "" else ""
            elif isinstance(raw_cost, (int, float)):
                normalized_cost = float(raw_cost)
        except Exception:
            pass
        # If this is a percent-of-base-rent expense, store as a fraction (e.g., 5 -> 0.05)
        if str(factor_text).lower() == "percent of base rent" and isinstance(normalized_cost, (int, float, float)):
            try:
                normalized_cost = normalized_cost / 100.0
            except Exception:
                pass
        row[8] = normalized_cost

        # Column J: Statistic
        base_summary_row = 6 + retail_income_length  # 'Total Base Retail Income' summary row
        # IF(G{row}="per SF / Yr.", $F${base_summary_row}, IF(G{row}="Percent of Base Rent", $L${base_summary_row}, ""))
        row[9] = f'=IF(G{current_row}="per SF / Yr.",$F${base_summary_row},IF(G{current_row}="Percent of Base Rent",$L${base_summary_row},""))'

        # Column K: Rent Type Included (text: Both, Gross, NNN, Neither)
        row[10] = expense.get("rent_type_included", "")

        # Column L: Annual = IF(J is blank, I, I*J)
        row[11] = f'=IF(J{current_row}="",I{current_row},I{current_row}*J{current_row})'

        # Monthly columns (M..EQ) grow the annual by month (use growth in E column)
        reference_occ_row = 6 + retail_income_length + 1
        for col_idx in range(13, 147):  # M..EQ
            # Convert column index to column letter(s)
            if col_idx < 26:
                col_letter = chr(ord('A') + col_idx)
            else:
                col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
            # Use Annual (L) and Growth (E)
            row[col_idx] = f"=IFERROR($L{current_row}/12*{col_letter}${reference_occ_row}*(1+$E{current_row})^(ROUNDUP({col_letter}$5/12,0)-1),0)"
        
        data_rows.append(row)
        # Per-row number formats for I (Expense) and J (Statistic) based on factor
        # Determine formats
        is_percent = factor_text.lower() == "percent of base rent"
        # I column format
        per_row_formats.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": current_row - 1,
                    "endRowIndex": current_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "PERCENT" if is_percent else "NUMBER",
                            "pattern": "0.00%" if is_percent else "$#,##0.00"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        })
        # J column format (Statistic): currency when percent-of-base, number otherwise
        per_row_formats.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": current_row - 1,
                    "endRowIndex": current_row,
                    "startColumnIndex": 9,  # J
                    "endColumnIndex": 10
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "NUMBER",
                            "pattern": "$#,##0" if is_percent else "#,##0"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat"
            }
        })
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{start_row}:EQ{start_row + retail_expenses_length - 1}",
        "values": data_rows
    }

    # Formatting for inserted rows: B–L non-bold; only column I blue; plus per-row number formats
    end_row = start_row + retail_expenses_length  # exclusive
    format_requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B
                    "endColumnIndex": 12   # up to L (exclusive)
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
                "fields": "userEnteredFormat.textFormat.bold"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": 8,  # I
                    "endColumnIndex": 9
                },
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        }
    ]
    # Ensure all inserted rows (full width A..EQ) are regular (not bold)
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row - 1,
                "endRowIndex": end_row,
                "startColumnIndex": 0,   # A
                "endColumnIndex": 147    # EQ is 147th column (0-based end exclusive)
            },
            "cell": {"userEnteredFormat": {"textFormat": {"bold": False}}},
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })
    # Append per-row number format requests
    format_requests.extend(per_row_formats)

    return insert_request, update_payload, format_requests

def get_retail_expenses_summary_row_industrial(retail_income, retail_expenses, sheet_name):
    """
    Creates the summary row for retail expenses.
    Row position: 17 + len(retail_income)*2 + len(retail_expenses)
    """
    retail_income_length = len(retail_income)
    retail_expenses_length = len(retail_expenses)
    
    # Calculate row position
    summary_row = 17 + retail_income_length * 2 + retail_expenses_length
    
    # Initialize row with empty strings
    row = [""] * 147
    
    # Column B: text
    row[1] = "Total Retail Operating Expenses"
    
    # Column I: =IFERROR(J{summary_row}/F{6 + retail_income_length},0)
    reference_row = 6 + retail_income_length
    row[8] = f"=IFERROR(J{summary_row}/F{reference_row},0)"
    
    # Column J: SUM of retail expenses column J
    expenses_start_row = 17 + retail_income_length * 2
    expenses_end_row = expenses_start_row + retail_expenses_length - 1
    row[11] = f"=SUM(L{expenses_start_row}:L{expenses_end_row})"
    
    # Columns L through EO (columns 11 through 144): SUM of expense rows for that column
    for col_idx in range(13, 147):  # L is column 11 (0-based), EO is column 144
        # Convert column index to column letter(s)
        if col_idx < 26:
            col_letter = chr(ord('A') + col_idx)
        else:
            col_letter = chr(ord('A') + col_idx // 26 - 1) + chr(ord('A') + col_idx % 26)
        row[col_idx] = f"=SUM({col_letter}{expenses_start_row}:{col_letter}{expenses_end_row})"
    
    # Update payload
    update_payload = {
        "range": f"'{sheet_name}'!A{summary_row}:EQ{summary_row}",
        "values": [row]
    }
    
    return update_payload

def insert_expense_rows_to_sheet_payloads(spreadsheet, sheet_name, expenses, model_variable_mapping):
    """
    Generates insert and update payloads for inserting rows into the given expense sheet.
    Returns insert_request and update_payloads instead of executing them.
    """
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]
    # print("EXPENSES", expenses)
    # Sort so that "Total percent of other expenses" is always at the end
    expenses = sorted(expenses, key=lambda x: x.get("factor") == "Total percent of other expenses")
    from string import ascii_uppercase

    def generate_excel_columns(start='J', end='EL'):
        def col_index(c):
            result = 0
            for i, char in enumerate(reversed(c)):
                result += (ascii_uppercase.index(char) + 1) * (26 ** i)
            return result

        def col_name(index):
            name = ""
            while index > 0:
                index, remainder = divmod(index - 1, 26)
                name = chr(65 + remainder) + name
            return name

        start_idx = col_index(start)
        end_idx = col_index(end)
        return [col_name(i) for i in range(start_idx, end_idx + 1)]

    month_columns = generate_excel_columns('J', 'EL')


    format_requests = []
    for idx, exp in enumerate(expenses):
        start_row = 1 + idx  # row 2 is index 1
        end_row = start_row + 1
        factor = exp.get("factor")

        # Cost column (C)
        # Determine the cost_per value for decimal logic
        cost_per_value = None
        try:
            # Try to get the value from exp.get("cost_per_value") or exp.get("cost_per")
            # Accept both numeric and string representations
            val = exp.get("cost_per_value", exp.get("cost_per"))
            if isinstance(val, (int, float)):
                cost_per_value = float(val)
            elif isinstance(val, str):
                # Remove any non-numeric characters (like $ or commas)
                import re
                num_str = re.sub(r"[^\d.]", "", val)
                if num_str:
                    cost_per_value = float(num_str)
        except Exception:
            cost_per_value = None

        if factor in ["Percent of Purchase Price", "Percent of Acquisition Loan", "Total percent of other expenses", "Percent of Property Taxes", "Percent of Insurance Cost"]:
            cost_number_format = {"type": "PERCENT", "pattern": "0.00%"}
        elif factor.lower() == "per Unit".lower():
            if cost_per_value is not None and cost_per_value >= 100:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0\"/unit\""}
            else:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00\"/unit\""}
        elif factor.lower() == "per SF".lower():
            if cost_per_value is not None and cost_per_value >= 100:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0\"/sf\""}
            else:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00\"/sf\""}
        elif factor.lower() == "per Month".lower():
            if cost_per_value is not None and cost_per_value >= 100:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0\"/month\""}
            else:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00\"/month\""}
        else:  # Total or anything else
            if cost_per_value is not None and cost_per_value >= 100:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0"}
            else:
                cost_number_format = {"type": "NUMBER", "pattern": "$#,##0.00"}

                
        format_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": 2,  # C
                    "endColumnIndex": 3
                },
                "cell": {"userEnteredFormat": {"numberFormat": cost_number_format}},
                "fields": "userEnteredFormat.numberFormat"
            }
        })

        # Ensure input values in columns B, C, and D use primary blue text color
        # Using a Google-like primary blue (#1E88E5)
        format_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": 1,  # B (inclusive)
                    "endColumnIndex": 4     # D (exclusive)
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {
                            "foregroundColor": {
                                "red": 0,
                                "green": 0,
                                "blue": 1
                            }
                        }
                    }
                },
                "fields": "userEnteredFormat.textFormat.foregroundColor"
            }
        })

        # Statistic column (D)
        if factor.lower() == "per Unit".lower():
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" units\""}
        elif factor.lower() == "per SF".lower():
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" sf\""}
        elif factor.lower() == "per Month".lower():
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0\" months\""}
        else:
            stat_number_format = {"type": "NUMBER", "pattern": "#,##0"}
        format_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": 3,  # D
                    "endColumnIndex": 4
                },
                "cell": {"userEnteredFormat": {"numberFormat": stat_number_format}},
                "fields": "userEnteredFormat.numberFormat"
            }
        })

    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,  # First inserted row (after header)
                "endRowIndex": 1 + len(expenses),  # Last inserted row + 1
                "startColumnIndex": 6,  # G
                "endColumnIndex": 8     # H (exclusive)
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "NUMBER",
                        "pattern": '"Month " #,##0'
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat"
        }
    })

    # Ensure all inserted rows are not bold (regular text) across the row
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,  # first inserted row (after header)
                "endRowIndex": 1 + len(expenses),  # through last inserted row
                "startColumnIndex": 0,  # from column A
                "endColumnIndex": 200   # up to column GR (~200), safely covering used columns
            },
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {
                        "bold": False
                    }
                }
            },
            "fields": "userEnteredFormat.textFormat.bold"
        }
    })

    

    # Step 1: Create insert request
    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": 1,
                "endIndex": 1 + len(expenses)
            },
            "inheritFromBefore": False
        }
    }

    # Step 2: Separate out "Total percent of other expenses"
    percent_expense = next((e for e in expenses if e["factor"] and e["factor"].lower() == "Total percent of other expenses".lower()), None)
    normal_expenses = [e for e in expenses if e["factor"] and e["factor"].lower() != "Total percent of other expenses".lower()]

    rows_to_insert = []


    purchase_price_location = get_mapped_cell_location(model_variable_mapping, 'General Property Assumptions', 'Acquisition Price')
    acquisition_loan_location = get_mapped_cell_location(model_variable_mapping, 'Other Reference', 'Max Loan Size')
    # Add normal expenses first
    for i, exp in enumerate(normal_expenses):
        row_index = i + 2  # 1-based index after header row

        name = exp.get("name", "")
        factor = exp.get("factor", "")
        cost_per = float(exp.get("cost_per") or 0)
        start_month = int(exp.get("start_month") or 0)
        end_month = int(exp.get("end_month") or 0)

        # Handle statistic
        statistic = "" 
        if factor.lower() == "Total".lower():
            statistic = ""
        elif factor.lower() == "Percent of Purchase Price".lower():
            statistic = f"={purchase_price_location}"
            cost_per = cost_per/100 
        elif factor.lower() == "Percent of Acquisition Loan".lower():
            statistic = f"={acquisition_loan_location}"
            cost_per = cost_per/100
        elif factor.lower() == "Percent of Property Taxes".lower():
            statistic = f"='Operating Expenses'!I2"
            cost_per = cost_per/100
        elif factor.lower() == "Percent of Insurance Cost".lower():
            statistic = f"='Operating Expenses'!I3"
            cost_per = cost_per/100
        else:
            statistic = float(exp.get("statistic") or 0)
            cost_per = cost_per

        # Total formula logic
        if factor.lower() == "Total".lower():
            total_formula = f"=C{row_index}"
        else:
            total_formula = f"=C{row_index}*D{row_index}"

        # elif factor == "Percent of Purchase Price":
        #     total_formula = f"=C{row_index}*{purchase_price_location}"
        #     cost_per = f"{cost_per}%"
        # elif factor == "Percent of Acquisition Loan":
        #     total_formula = f"=C{row_index}*{acquisition_loan_location}"
        #     cost_per = f"{cost_per}%"
        # else:
        #     total_formula = ""

        
        row = [name, factor, cost_per, statistic, total_formula, "", start_month, end_month, ""]

        # Monthly formulas from J to EL
        for month_offset, col_letter in enumerate(month_columns):
            formula = f"=IF(AND($G{row_index}={col_letter}$1,$H{row_index}={col_letter}$1),$E{row_index},IF($G{row_index}>{col_letter}$1,0,IF($H{row_index}<{col_letter}$1,0,$E{row_index}/($H{row_index}-$G{row_index}+1))))"
            row.append(formula)

        rows_to_insert.append(row)

    # Add "Total percent of other expenses" as the last row (if it exists)
    if percent_expense:
        percent_row_index = len(rows_to_insert) + 2  # Inserted after the rest
        name = percent_expense.get("name", "")
        cost_per_val = float(percent_expense.get("cost_per") or 0)
        cost_per = f"{cost_per_val}%"  # Add percent sign

        start_month = int(percent_expense.get("start_month") or 0)
        end_month = int(percent_expense.get("end_month") or 0)

        # Statistic = sum of previous total column; if no prior rows, use 0 to avoid SUM(E2:E1)
        statistic_formula = "0" if len(rows_to_insert) == 0 else f"=SUM(E2:E{percent_row_index - 1})"
        total_formula = f"=C{percent_row_index}*D{percent_row_index}"
        # if expenses.length > 1:
        #     total_formula = f"=C{percent_row_index}*D{percent_row_index}"
        # else: 
        #     total_formula = f"0"

        row = [name, "Total percent of other expenses", cost_per, statistic_formula, total_formula, "", start_month, end_month, ""]

        for col_letter in month_columns:
            formula = f"=IF(AND($G{percent_row_index}={col_letter}$1,$H{percent_row_index}={col_letter}$1),$E{percent_row_index},IF($G{percent_row_index}>{col_letter}$1,0,IF($H{percent_row_index}<{col_letter}$1,0,$E{percent_row_index}/($H{percent_row_index}-$G{percent_row_index}+1))))"
            row.append(formula)

        rows_to_insert.append(row)

    # Step 3: Create main data update payload
    start_row = 2
    end_row = 1 + len(rows_to_insert)
    end_col_letter = "EL"
    range_str = f"'{sheet_name}'!A{start_row}:{end_col_letter}{end_row}"
    
    main_update_payload = {
        "range": range_str,
        "values": rows_to_insert
    }

    # Step 4: Create Total row update payloads
    total_row_index = end_row + 1

    # Update E (total column)
    total_formula_value = f"=SUM(E2:E{end_row})"
    total_column_update_payload = {
        "range": f"{sheet_name}!E{total_row_index}",
        "values": [[total_formula_value]]
    }

    # Update each month column (J to EL) with a SUM formula
    month_sums = []
    for col_letter in month_columns:
        formula = f"=SUM({col_letter}2:{col_letter}{end_row})"
        month_sums.append(formula)

    # Create month totals update payload
    month_totals_update_payload = {
        "range": f"{sheet_name}!J{total_row_index}:{month_columns[-1]}{total_row_index}",
        "values": [month_sums]
    }

    return insert_request, [main_update_payload, total_column_update_payload, month_totals_update_payload], format_requests



def run_full_sheet_update(
    spreadsheet,
    market_json,
    rental_assumptions_json,
    rental_growth_json,
    amenity_income_json,
    operating_expenses_json, 
    model_variable_mapping,
    address,
    retail_income,
    expenses_json,
    property_name
):

    industrial_model = False
    if len(operating_expenses_json) == 0:
        industrial_model = True

    retail_expenses = [expense for expense in expenses_json if expense.get('type') == 'Retail']
    
    print("[run_full_sheet_update] START")
    try:
        print(f"[run_full_sheet_update] Input sizes -> market:{len(market_json) if market_json is not None else 'None'}, "
              f"rental_assumptions:{len(rental_assumptions_json) if rental_assumptions_json is not None else 'None'}, "
              f"rental_growth:{len(rental_growth_json) if rental_growth_json is not None else 'None'}, "
              f"amenity_income:{len(amenity_income_json) if amenity_income_json is not None else 'None'}, "
              f"operating_expenses:{len(operating_expenses_json) if operating_expenses_json is not None else 'None'}, "
              f"retail_income:{len(retail_income) if retail_income is not None else 'None'}, "
              f"expenses:{len(expenses_json) if expenses_json is not None else 'None'}")
    except Exception as _e:
        print("[run_full_sheet_update] Failed to print input sizes (some inputs may be None)")
    # === Sheet handles ===
    if len(market_json) > 0 and len(rental_assumptions_json) > 0:
        print("[run_full_sheet_update] Fetching worksheets: Market Rent Assumptions, Rental Assumptions, Rent Roll Model")
        market_ws = spreadsheet.worksheet("Market Rent Assumptions")
        rental_ws = spreadsheet.worksheet("Rental Assumptions")
        rent_ws = spreadsheet.worksheet("Rent Roll Model")
        rental_growth_ws = rent_ws
        rent_roll_ws = rent_ws
    else:
        print("[run_full_sheet_update] Skipping rental/market worksheets (no market_json or rental_assumptions_json)")

    assumptions_ws = spreadsheet.worksheet("Assumptions")
    print("[run_full_sheet_update] Loaded worksheet: Assumptions")

    # === Insert & Value Data Preparation ===
    if (len(market_json) > 0 and len(rental_assumptions_json) > 0):

        print(f"[run_full_sheet_update] Preparing market insert ops with market:{len(market_json)}, rentals:{len(rental_assumptions_json)}")
        market_insert_request, market_value_data, market_format_request = get_market_rent_insert_ops(market_ws, market_json, rental_assumptions_json)
        print(f"[run_full_sheet_update] market_insert_request:{len(market_insert_request)} value_data:{len(market_value_data)} format_reqs:{len(market_format_request)}")

        rental_requests, rental_value_data, rental_total_row = get_rental_assumptions_insert_ops(
        rental_ws, rental_assumptions_json, market_start_row=2, market_end_row=2 + len(market_json) - 1
        )
        print(f"[run_full_sheet_update] rental_requests:{len(rental_requests)} rental_value_data:{len(rental_value_data)} rental_total_row:{len(rental_total_row)}")

        assumptions_values, assumptions_format, assumption_row_mapping = get_rental_growth_assumptions_inserts(
            assumptions_ws, rental_growth_json, model_variable_mapping
        )
        print(f"[run_full_sheet_update] assumptions_values:{len(assumptions_values)} assumptions_format:{len(assumptions_format)} assumption_row_mapping:{len(assumption_row_mapping)}")

        rent_roll_inserts, rent_roll_values = get_rent_roll_growth_inserts(
            rent_ws, rental_growth_json, assumption_row_mapping, rent_start_row=3
        )
        print(f"[run_full_sheet_update] rent_roll_inserts:{len(rent_roll_inserts)} rent_roll_values_rows:{len(rent_roll_values)}")
    else: 

        print("[run_full_sheet_update] No market/rental data; initializing empty requests for those sections")
        market_insert_request = []
        market_value_data = []
        market_format_request = []
        rental_requests = []
        rental_value_data = []
        rental_total_row = []
        assumptions_values = []
        assumptions_format = []
        assumption_row_mapping = []
        rent_roll_inserts = []
        rent_roll_values = []


    print(f"[run_full_sheet_update] Preparing amenity income inserts: items:{len(amenity_income_json)}")
    amenity_income_insert_request = get_amenity_income_insert_request(
        spreadsheet, amenity_income_json, start_row=2, sheet_name="Amenity Income"
    )
    print(f"[run_full_sheet_update] amenity_income_insert_request ops:{len(amenity_income_insert_request)}")

    noi_walk_amenity_start_row = 25
    noi_amenity_start_row = 10
    if len(operating_expenses_json) > 0:
        
        noi_walk_amenity_start_row = 25
        noi_amenity_start_row = 10
    else: 
        ## industrial model
        noi_walk_amenity_start_row = 20
        noi_amenity_start_row = 8

    amenity_income_insert_request_noi = get_amenity_income_insert_request_noi(
        spreadsheet, amenity_income_json, start_row=noi_walk_amenity_start_row, sheet_name="NOI Walk"
    )
    print(f"[run_full_sheet_update] amenity_income_insert_request_noi ops:{len(amenity_income_insert_request_noi)}")

    amenity_noi_summary_insert_request = get_noi_summary_insert_request(
        spreadsheet, len(amenity_income_json), start_row=noi_amenity_start_row, sheet_name="NOI"
    )

    print(f"[run_full_sheet_update] amenity_noi_summary_insert_request ops:{len(amenity_noi_summary_insert_request)}")

    if len(operating_expenses_json) > 0:

        print(f"[run_full_sheet_update] Preparing operating expenses inserts: items:{len(operating_expenses_json)}")
        operating_expenses_insert = get_operating_expenses_insert_request(
            spreadsheet, operating_expenses_json, start_row=2, sheet_name="Operating Expenses"
        )
        print(f"[run_full_sheet_update] operating_expenses_insert ops:{len(operating_expenses_insert)}")

        operating_expense_rows_insert = get_operating_expense_row_insert_request(
            spreadsheet, operating_expenses_json, amenity_income_json, sheet_name="NOI Walk", start_base_row=30
        )
        print(f"[run_full_sheet_update] operating_expense_rows_insert ops:{len(operating_expense_rows_insert)}")
    else: 
        print(f"[run_full_sheet_update] Preparing operating expenses inserts: items:{len(operating_expenses_json)}")
        operating_expenses_insert = []

        operating_expense_rows_insert = get_operating_expense_row_insert_request_industrial(
            spreadsheet, retail_expenses, amenity_income_json, sheet_name="NOI Walk", start_base_row=23
        )
        print(f"[run_full_sheet_update] operating_expense_rows_insert ops:{len(operating_expense_rows_insert)}")

    
    print("assumption_row_mapping", assumption_row_mapping)
    if len(assumption_row_mapping) > 0:
        inflation_updates = get_inflation_reference_updates(assumption_row_mapping)
    else: 
        inflation_updates = []


    if industrial_model:
        noi_expense_insert, noi_expense_update, noi_expense_reset_format = get_noi_expense_rows_insert_and_update_industrial(
            spreadsheet, retail_expenses, amenity_income_json
        )
    else: 

        noi_expense_insert, noi_expense_update, noi_expense_reset_format = get_noi_expense_rows_insert_and_update(
            spreadsheet, operating_expenses_json, amenity_income_json
        )
    print(f"[run_full_sheet_update] NOI expense inserts:{len(noi_expense_insert)} updates:{len(noi_expense_update)} reset_format:{len(noi_expense_reset_format)}")

    if len(rental_assumptions_json) > 0:

        print(f"[run_full_sheet_update] Building rent roll assumption inserts for {len(rental_assumptions_json)} rentals")
        rent_roll_assumptions_insert, rent_roll_output_values = get_rent_roll_assumption_row_inserts(
            rent_roll_ws, rental_assumptions_json, [x for x in rental_growth_json if x.get("type") == "rental"], model_variable_mapping, start_row=7
        )
        print(f"[run_full_sheet_update] rent_roll_assumptions_insert ops:{len(rent_roll_assumptions_insert)} rent_roll_output_values rows:{len(rent_roll_output_values)}")

        growth_formula_update, growth_format_request = get_rent_roll_growth_formula_updates(
            rental_growth_ws, [x for x in rental_growth_json if x.get("type") == "rental"]
        )
        print(f"[run_full_sheet_update] growth_formula_update ops:{len(growth_formula_update)} growth_format_request ops:{len(growth_format_request)}")

        vacancy_insert_request, vacancy_formula_update, vacancy_format_request = get_rent_roll_vacancy_row_inserts_and_updates(
            rent_roll_ws, rental_assumptions_json, [x for x in rental_growth_json if x.get("type") == "rental"]
        )
        print(f"[run_full_sheet_update] vacancy_insert_request ops:{len(vacancy_insert_request)} vacancy_formula_update ops:{len(vacancy_formula_update)} vacancy_format_request ops:{len(vacancy_format_request)}")


        # === Update Payloads ===
        market_formula_update = get_market_rent_formula_updates(
            market_json, rental_start_row=2, rental_end_row=2 + len(rental_assumptions_json) - 1, market_start_row=2
        )
        print(f"[run_full_sheet_update] market_formula_update ops:{len(market_formula_update)}")
        noi_growth_formula_update = get_noi_growth_factor_formula_update(
            assumption_row_mapping,
            rental_growth_json,
            noi_formula_row=9,
            noi_formula_col_start=5,
            noi_formula_col_end=132,
            sheet_name="NOI Walk"
        )
        print(f"[run_full_sheet_update] noi_growth_formula_update ops:{len(noi_growth_formula_update)}")

        total_summary_update, total_summary_format = get_total_summary_row_updates(rent_roll_ws,
            sheet_name="Rent Roll Model", start_row=7 + len([entry for entry in rental_growth_json if entry.get("type") == "rental"]), num_rows=len(rental_assumptions_json), model_variable_mapping=model_variable_mapping
        )
        print(f"[run_full_sheet_update] total_summary_update ops:{len(total_summary_update)} total_summary_format ops:{len(total_summary_format)}")

    else: 
        print("[run_full_sheet_update] No rental_assumptions_json; skipping rent roll & related formula updates")
        rent_roll_assumptions_insert = []
        rent_roll_output_values = []
        growth_formula_update = []
        growth_format_request = []
        vacancy_insert_request = []
        vacancy_formula_update = []
        vacancy_format_request = []
        market_formula_update = []
        noi_growth_formula_update = []
        total_summary_update = []
        total_summary_format = []





    if len(amenity_income_json) > 0:
        amenity_income_update = get_amenity_income_update_payload(
            sheet_name="Amenity Income", amenity_income_json=amenity_income_json, start_row=2
        )
        amenity_income_totals_update = get_amenity_income_totals_update_payload(
            sheet_name="Amenity Income", start_row=2, num_amenities=len(amenity_income_json)
        )
        amenity_income_format = get_amenity_income_format_requests(
            spreadsheet, sheet_name="Amenity Income", start_row=2, num_amenities=len(amenity_income_json)
        )
        if industrial_model:
            amenity_income_formula_update = get_amenity_income_formula_update_industrial(
                amenity_json=amenity_income_json, start_row=noi_walk_amenity_start_row, start_col=5, num_months=132
            )
        else:
            amenity_income_formula_update = get_amenity_income_formula_update(
                amenity_json=amenity_income_json, start_row=noi_walk_amenity_start_row, start_col=5, num_months=132
            )
        noi_summary_update = get_noi_summary_row_update_payload(
            num_rows=len(amenity_income_json), walk_start_row=noi_walk_amenity_start_row, noi_start_row=noi_amenity_start_row
    )
        print(f"[run_full_sheet_update] amenity_income_update ops:{len(amenity_income_update)} amenity_income_totals_update ops:{len(amenity_income_totals_update)} "
              f"amenity_income_format ops:{len(amenity_income_format)} amenity_income_formula_update ops:{len(amenity_income_formula_update)} "
              f"noi_summary_update ops:{len(noi_summary_update)}")

    else:
        amenity_income_update = []
        amenity_income_totals_update = []
        amenity_income_format = []
        amenity_income_formula_update = []
        noi_summary_update = []
        print("[run_full_sheet_update] No amenity_income_json; skipping amenity updates")
    

    if len(operating_expenses_json) > 0:
        operating_expenses_update = get_operating_expenses_update_payload(
            operating_expenses_json, rental_assumptions_json, amenity_income_json, model_variable_mapping, sheet_name="Operating Expenses", start_row=2
        )

        operating_expenses_format = get_operating_expenses_format_payload(
            operating_expenses_json, spreadsheet, sheet_name="Operating Expenses", start_row=2
        )

        operating_expenses_sum_row_update = get_operating_expenses_sum_row_payload(
            operating_expenses_json, sheet_name="Operating Expenses", start_row=2
        )

        operating_expense_formula_updates = get_operating_expense_formula_update_payloads(
            operating_expenses_json, amenity_income_json, noi_sheet="NOI Walk",
            op_exp_sheet="Operating Expenses", start_base_row=30, start_col=5, num_months=132
        )
        print(f"[run_full_sheet_update] operating_expenses_update ops:{len(operating_expenses_update)} "
              f"operating_expenses_format ops:{len(operating_expenses_format)} "
              f"operating_expenses_sum_row_update ops:{len(operating_expenses_sum_row_update)} "
              f"operating_expense_formula_updates ops:{len(operating_expense_formula_updates)}")

        expense_sum_row_to_noi_walk = get_expense_sum_row_to_noi_walk_payload(
        amenity_income_json, operating_expenses_json
        )
        print(f"[run_full_sheet_update] expense_sum_row_to_noi_walk ops:{len(expense_sum_row_to_noi_walk)}")

    else: 
        operating_expenses_update = []
        operating_expenses_format = []
        operating_expenses_sum_row_update = []
        operating_expense_formula_updates = get_operating_expense_formula_update_payloads_industrial(
            retail_expenses,
            amenity_income_json,
            noi_sheet="NOI Walk",
            op_exp_sheet="Retail Assumptions",
            start_base_row=23,
            start_col=5,
            num_months=132,
            ra_start_row=17 + len(retail_income) * 2
        )
        expense_sum_row_to_noi_walk = get_expense_sum_row_to_noi_walk_payload_industrial(
        amenity_income_json, retail_expenses
        )
        print("[run_full_sheet_update] No operating_expenses_json; skipping op-exp updates")



    try:
        re_stabilization_update = get_restabilization_update_payload(model_variable_mapping, rental_assumptions_json)
    except Exception:
        re_stabilization_update = []
    print(f"[run_full_sheet_update] re_stabilization_update ops:{len(re_stabilization_update)}")

    
    address_update = get_address_update_payload(address, model_variable_mapping)
    if property_name:
        property_name_update = get_property_name_update_payload(property_name, model_variable_mapping)
    else:
        property_name_update = []


    if industrial_model:
        # =COUNTA(UNIQUE('Retail Assumptions'!$B$6:B9))
        number_of_spaces_update = get_number_of_spaces_update_payload(model_variable_mapping, retail_income, sheet_name="Retail Assumptions")

        # =+SUMIFS('Retail Assumptions'!$L$6:$L$9,'Retail Assumptions'!$G$6:$G$9,">="&1)/SUMIFS('Retail Assumptions'!$F$6:$F$9,'Retail Assumptions'!$G$6:$G$9,">="&1)
        in_place_rent_update = get_in_place_rent_update_payload(model_variable_mapping, retail_income, sheet_name="Retail Assumptions")
    else:
        number_of_spaces_update = []
        in_place_rent_update = []

   
    ntm_formula_update = get_ntm_update_payload(amenity_income_json, operating_expenses_json)
    print(f"[run_full_sheet_update] ntm_formula_update ops:{len(ntm_formula_update)}")


    retail_rates = [rate for rate in rental_growth_json if rate.get('type') == 'retail']
    retail_growth_rate = retail_rates[0]['value'] if retail_rates else 0
    print(f"[run_full_sheet_update] retail_rates_count:{len(retail_rates)} retail_growth_rate:{retail_growth_rate}")

    
    print(f"[run_full_sheet_update] retail_expenses_count:{len(retail_expenses)}")

    if len(rental_assumptions_json) > 0:

        vacancy_sum_update = get_rent_roll_vacancy_sum_row_update(rent_roll_ws, rental_assumptions_json, rental_growth_json)
        print(f"[run_full_sheet_update] vacancy_sum_update ops:{len(vacancy_sum_update)}")


    else: 
        vacancy_sum_update = []
        print("[run_full_sheet_update] No rental_assumptions_json; vacancy_sum_update skipped")


    if len(operating_expenses_json) > 0:
        noi_egi_update = get_noi_walk_egi_row_update_payload(amenity_income_json, amenity_start_row=noi_walk_amenity_start_row, target_base_row=noi_walk_amenity_start_row + 2)
    else:
        noi_egi_update = get_noi_walk_egi_row_update_payload_industrial(amenity_income_json, amenity_start_row=noi_walk_amenity_start_row, target_base_row=noi_walk_amenity_start_row)


    if len(retail_income) > 0:
        if industrial_model:
            retail_assumptions_insert_request, retail_assumptions_update_payload, retail_assumptions_format_requests = get_retail_assumptions_inserts_industrial(spreadsheet, retail_income, 'Retail Assumptions')
            retail_assumptions_summary_update_payload = get_retail_assumptions_summary_row_industrial(retail_income, 'Retail Assumptions')
            retail_assumptions_occ_update_payload = get_retail_assumptions_occ_row_industrial(retail_income, 'Retail Assumptions')
            retail_recovery_insert_request, retail_recovery_update_payload, retail_recovery_format_requests = get_retail_recovery_inserts_industrial(spreadsheet, retail_income, retail_expenses, 'Retail Assumptions')
            retail_recovery_summary_update_payload = get_retail_recovery_summary_row_industrial(retail_income, 'Retail Assumptions')
            retail_expenses_insert_request, retail_expenses_update_payload, retail_expenses_format_requests = get_retail_expenses_inserts_industrial(spreadsheet, retail_income, retail_expenses, retail_growth_rate, 'Retail Assumptions', model_variable_mapping)
            retail_expenses_summary_update_payload = get_retail_expenses_summary_row_industrial(retail_income, retail_expenses, 'Retail Assumptions')
        else: 
            retail_assumptions_insert_request, retail_assumptions_update_payload, retail_assumptions_format_requests = get_retail_assumptions_inserts(spreadsheet, retail_income, 'Retail Assumptions')
            retail_assumptions_summary_update_payload = get_retail_assumptions_summary_row(retail_income, 'Retail Assumptions')
            retail_assumptions_occ_update_payload = get_retail_assumptions_occ_row(retail_income, 'Retail Assumptions')
            retail_recovery_insert_request, retail_recovery_update_payload, retail_recovery_format_requests = get_retail_recovery_inserts(spreadsheet, retail_income, retail_expenses, 'Retail Assumptions')
            retail_recovery_summary_update_payload = get_retail_recovery_summary_row(retail_income, 'Retail Assumptions')
            retail_expenses_insert_request, retail_expenses_update_payload, retail_expenses_format_requests = get_retail_expenses_inserts(spreadsheet, retail_income, retail_expenses, retail_growth_rate, 'Retail Assumptions')
            retail_expenses_summary_update_payload = get_retail_expenses_summary_row(retail_income, retail_expenses, 'Retail Assumptions')
    else:
        retail_assumptions_insert_request, retail_assumptions_update_payload, retail_assumptions_format_requests = [], [], []
        retail_assumptions_summary_update_payload = []
        retail_assumptions_occ_update_payload = []
        retail_recovery_insert_request, retail_recovery_update_payload, retail_recovery_format_requests = [], [], []
        retail_recovery_summary_update_payload = []
        retail_expenses_insert_request, retail_expenses_update_payload, retail_expenses_format_requests = [], [], []
        retail_expenses_summary_update_payload = []


    closing_costs_filtered = [expense for expense in expenses_json if expense.get('type') == 'Closing Costs']
    closing_cost_insert_request, closing_cost_update_payloads, closing_cost_format_requests = insert_expense_rows_to_sheet_payloads(spreadsheet, 'Closing Costs', closing_costs_filtered, model_variable_mapping)
    hard_costs_filtered = [expense for expense in expenses_json if expense.get('type') == 'Hard Costs']
    hard_costs_insert_request, hard_costs_update_payloads, hard_costs_format_requests = insert_expense_rows_to_sheet_payloads(spreadsheet, 'Hard Costs', hard_costs_filtered, model_variable_mapping)
    legal_costs_filtered = [expense for expense in expenses_json if expense.get('type') == 'Legal and Pre-Development Costs']
    legal_costs_insert_request, legal_costs_update_payloads, legal_costs_format_requests = insert_expense_rows_to_sheet_payloads(spreadsheet, 'Legal and Pre-Development Costs', legal_costs_filtered, model_variable_mapping)
    reserves_filtered = [expense for expense in expenses_json if expense.get('type') == 'Reserves']
    reserves_insert_request, reserves_update_payloads, reserves_format_requests = insert_expense_rows_to_sheet_payloads(spreadsheet, 'Reserves', reserves_filtered, model_variable_mapping)
    # === Combine Insert & Format Requests ===
    insert_requests = [
        market_insert_request,
        market_format_request,
        *rental_requests,
        assumptions_format,
        amenity_income_insert_request,
        *amenity_income_format,
        rent_roll_inserts,
        amenity_income_insert_request_noi,
        amenity_noi_summary_insert_request,
        operating_expenses_insert,
        operating_expense_rows_insert,
        noi_expense_insert,
        rent_roll_assumptions_insert,
        vacancy_insert_request,
        *([] if len(retail_income) == 0 else [retail_assumptions_insert_request]),
        *([] if len(retail_income) == 0 else [retail_recovery_insert_request]),
        *([] if len(retail_expenses) == 0 else [retail_expenses_insert_request]),
        closing_cost_insert_request,
        hard_costs_insert_request,
        legal_costs_insert_request,
        reserves_insert_request,
        amenity_income_format,
        noi_expense_reset_format,
        retail_assumptions_format_requests,
        retail_recovery_format_requests,
        retail_expenses_format_requests
    ]

    print("ALL INSERT REQUESTS", insert_requests)
    format_requests = [growth_format_request, vacancy_format_request, total_summary_format, *closing_cost_format_requests, *hard_costs_format_requests, *legal_costs_format_requests, *reserves_format_requests, *operating_expenses_format]

    # === Combine Update Payloads ===
    update_payloads = [
        market_value_data,
        rental_value_data,
        rental_total_row,
        assumptions_values,
        rent_roll_values,
        rent_roll_output_values,
        inflation_updates,
        noi_expense_update,
        market_formula_update,
        noi_growth_formula_update,
        growth_formula_update,
        vacancy_formula_update,
        vacancy_sum_update,
        total_summary_update,
        amenity_income_update,
        amenity_income_totals_update,
        amenity_income_formula_update,
        noi_summary_update,
        noi_egi_update,
        operating_expenses_update,
        operating_expenses_sum_row_update,
        *operating_expense_formula_updates,
        expense_sum_row_to_noi_walk,
        ntm_formula_update,
        re_stabilization_update,
        address_update,
        property_name_update,
        number_of_spaces_update,
        in_place_rent_update,
        *([] if len(retail_income) == 0 else [retail_assumptions_update_payload]), 
        *([] if len(retail_income) == 0 else [retail_assumptions_summary_update_payload]), 
        *([] if len(retail_income) == 0 else [retail_assumptions_occ_update_payload]), 
        *([] if len(retail_income) == 0 else [retail_recovery_update_payload]),
        *([] if len(retail_income) == 0 else [retail_recovery_summary_update_payload]),
        *([] if len(retail_expenses) == 0 else [retail_expenses_update_payload]),
        *([] if len(retail_expenses) == 0 else [retail_expenses_summary_update_payload]),
        closing_cost_update_payloads,
        hard_costs_update_payloads,
        legal_costs_update_payloads,
        reserves_update_payloads
    ]

    # === Run all operations ===
    spreadsheet.batch_update({"requests": insert_requests + format_requests})
    # spreadsheet.batch_update({"requests": insert_requests})

    spreadsheet.values_batch_update({"valueInputOption": "USER_ENTERED", "data": update_payloads})
    print("✅ All inserts and updates applied.")


def update_google_sheet_and_get_values(
    copied_sheet_id,
    copied_sheet_url,
    mapped_values,
    market_json,
    rental_assumptions_json,
    rental_growth_json,
    amenity_income_json,
    expenses_json,
    operating_expenses_json,
    retail_income_json
):
    print("📤 Starting Google Sheet generation workflow...")
    timings = {}

    sheets_service = build("sheets", "v4", credentials=creds)

    # Step 1: Fetch all mappings
    t0 = time.time()
    print("📊 Fetching mapping sheets...")
    result = sheets_service.spreadsheets().values().batchGet(
        spreadsheetId=copied_sheet_id,
        ranges=[
            "'Model Variable Mapping'",
            "'Table Mapping'",
            "'Variable Mapping'"
        ],
        valueRenderOption='FORMULA'
    ).execute()
    # print("RESULT", result)
    model_mapping_values = result['valueRanges'][0].get('values', [])
    # For all rows with length 3, append two empty strings
    model_mapping_values = [row + ['', ''] if len(row) == 3 else row for row in model_mapping_values]
    table_mapping_values = result['valueRanges'][1].get('values', [])
    variable_mapping_values = result['valueRanges'][2].get('values', [])

    df = pd.DataFrame(model_mapping_values[1:], columns=model_mapping_values[0])
    # print("DF", df)
    table_mapping_data = [dict(zip(table_mapping_values[0], row)) for row in table_mapping_values[1:]]
    variable_data = [dict(zip(variable_mapping_values[0], row)) for row in variable_mapping_values[1:]]

    t1 = time.time()
    timings['load_mapping'] = t1 - t0
    print(f"✅ Loaded mapping DataFrame with {len(df)} rows in {timings['load_mapping']:.3f}s")

    # Step 2: Update cell-level values
    t2 = time.time()
    update_copied_sheet_values(copied_sheet_id, mapped_values, df)
    t3 = time.time()
    timings['update_sheet'] = t3 - t2
    print(f"✅ Sheet update complete in {timings['update_sheet']:.3f}s")

    # ✅ Step 2.5: Run full structured insert/update
    print("🚀 Running full sheet data inserts/formulas...")
    spreadsheet = gs_client.open_by_key(copied_sheet_id)
    run_full_sheet_update(
        spreadsheet,
        market_json=market_json,
        rental_assumptions_json=rental_assumptions_json,
        rental_growth_json=rental_growth_json,
        amenity_income_json=amenity_income_json,
        operating_expenses_json=operating_expenses_json,
        model_variable_mapping=df,
        address="",
        retail_income=retail_income_json, 
        expenses_json=expenses_json,
        property_name=''
    )

    # Step 3: Extract tables
    t4 = time.time()
    tables = extract_tables_for_storage_batch(copied_sheet_id, table_mapping_data, sheets_service)
    t5 = time.time()
    timings['extract_tables'] = t5 - t4
    print(f"✅ Extracted {len(tables)} tables in {timings['extract_tables']:.3f}s")

    # Step 4: Extract variables
    t6 = time.time()
    variables = extract_variables_from_sheet_batch(copied_sheet_id, variable_data, sheets_service)
    t7 = time.time()
    timings['extract_variables'] = t7 - t6
    print(f"📈 Extracted variables: {variables} in {timings['extract_variables']:.3f}s")

    # Total time
    timings['total'] = t7 - t0
    print(f"⏱️ Total time: {timings['total']:.3f}s")
    print("🎯 Finished sheet generation and data extraction.")
    print(f"⏱️ Timings: {timings}")



    return {
        "sheet_url": copied_sheet_url,
        "levered_irr": variables.get("Levered IRR"),
        "levered_moic": variables.get("Levered MOIC"),
        "variables": variables,
        "tables": tables,
        "timings": timings
    }


def update_google_sheet_and_get_values_intermediate(
    copied_sheet_id,
    copied_sheet_url,
    mapped_values,
    market_json,
    rental_assumptions_json,
    rental_growth_json,
    amenity_income_json,
    expenses_json,
    operating_expenses_json,
    retail_income_json,
    address,
    property_name
):
    print("📤 Starting Google Sheet generation workflow...")
    timings = {}

    sheets_service = build("sheets", "v4", credentials=creds)

    # Step 1: Fetch all mappings
    t0 = time.time()
    print("📊 Fetching mapping sheets...")
    result = sheets_service.spreadsheets().values().batchGet(
        spreadsheetId=copied_sheet_id,
        ranges=[
            "'Model Variable Mapping'",
            "'Variable Mapping'"
        ],
        valueRenderOption='FORMULA'
    ).execute()
    model_mapping_values = result['valueRanges'][0].get('values', [])
    # For all rows with length 3, append two empty strings
    model_mapping_values = [row + ['', ''] if len(row) == 3 else row for row in model_mapping_values]
    variable_mapping_values = result['valueRanges'][1].get('values', [])

    df = pd.DataFrame(model_mapping_values[1:], columns=model_mapping_values[0])
    variable_data = [dict(zip(variable_mapping_values[0], row)) for row in variable_mapping_values[1:]]

    t1 = time.time()
    timings['load_mapping'] = t1 - t0
    print(f"✅ Loaded mapping DataFrame with {len(df)} rows in {timings['load_mapping']:.3f}s")

    print("MAPPED VALUES", mapped_values)
    # mapped_values = [x for x in mapped_values if x.get("section") == "General Property Assumptions"]
    # Step 2: Update cell-level values
    t2 = time.time()
    update_copied_sheet_values(copied_sheet_id, mapped_values, df)
    t3 = time.time()
    timings['update_sheet'] = t3 - t2
    print(f"✅ Sheet update complete in {timings['update_sheet']:.3f}s")
    t4 = time.time()
    # ✅ Step 2.5: Run full structured insert/update
    print("🚀 Running full sheet data inserts/formulas...")
    spreadsheet = gs_client.open_by_key(copied_sheet_id)
    run_full_sheet_update(
        spreadsheet,
        market_json=market_json,
        rental_assumptions_json=rental_assumptions_json,
        rental_growth_json=rental_growth_json,
        amenity_income_json=amenity_income_json,
        expenses_json=expenses_json,
        model_variable_mapping=df,
        address=address,
        retail_income=retail_income_json, 
        operating_expenses_json=operating_expenses_json,
        property_name=property_name
    )
    t5 = time.time()
    timings['run_full_sheet_update'] = t5 - t4
    print(f"✅ Full sheet update complete in {timings['run_full_sheet_update']:.3f}s")
    # Step 4: Extract variables
    t6 = time.time()
    variables = extract_variables_from_sheet_batch(copied_sheet_id, variable_data, sheets_service)
    t7 = time.time()
    timings['extract_variables'] = t7 - t6
    print(f"📈 Extracted variables: {variables} in {timings['extract_variables']:.3f}s")

    # Total time
    timings['total'] = t7 - t0
    print(f"⏱️ Total time: {timings['total']:.3f}s")
    print("🎯 Finished sheet generation and data extraction.")
    print(f"⏱️ Timings: {timings}")

    ## running again cuz growth rates... 
    result = sheets_service.spreadsheets().values().batchGet(
    spreadsheetId=copied_sheet_id,
    ranges=[
        "'Model Variable Mapping'",
        "'Variable Mapping'"
    ],
    valueRenderOption='FORMULA'
    ).execute()
    model_mapping_values = result['valueRanges'][0].get('values', [])
    # For all rows with length 3, append two empty strings
    model_mapping_values = [row + ['', ''] if len(row) == 3 else row for row in model_mapping_values]
    variable_mapping_values = result['valueRanges'][1].get('values', [])

    df = pd.DataFrame(model_mapping_values[1:], columns=model_mapping_values[0])

    # Fetch NOI table (entire sheet values)
    try:
        noi_resp = sheets_service.spreadsheets().values().get(
            spreadsheetId=copied_sheet_id,
            range="'NOI'"
        ).execute()
        noi_values = noi_resp.get("values", [])
    except Exception as e:
        print(f"⚠️ Failed to fetch NOI sheet values: {e}")
        noi_values = []

    return {
        "sheet_url": copied_sheet_url,
        "levered_irr": variables.get("Levered IRR"),
        "levered_moic": variables.get("Levered MOIC"),
        "variables": variables,
        "model_mapping": df.to_dict(orient="records"),
        "variable_mapping": variable_data,
        "timings": timings,
        "NOI": noi_values,
    }



def update_google_sheet_field_values_and_get_values(
    copied_sheet_id,
    mapped_values,
    model_mapping,
    variable_mapping
):
    print("📤 Starting Google Sheet generation workflow...")
    timings = {}

    # Convert model_mapping (list of dicts) back to DataFrame
    model_mapping_df = pd.DataFrame(model_mapping)
    sheets_service = build("sheets", "v4", credentials=creds)

    df = model_mapping_df
    variable_data = variable_mapping

    # Step 2: Update cell-level values
    t2 = time.time()
    update_copied_sheet_values(copied_sheet_id, mapped_values, df)
    t3 = time.time()
    timings['update_sheet'] = t3 - t2
    print(f"✅ Sheet update complete in {timings['update_sheet']:.3f}s")
    # ✅ Step 2.5: Run full structured insert/update
    print("🚀 Running full sheet data inserts/formulas...")
    # Step 4: Extract variables
    t6 = time.time()
    variables = extract_variables_from_sheet_batch(copied_sheet_id, variable_data, sheets_service)
    t7 = time.time()
    timings['extract_variables'] = t7 - t6
    # Fetch NOI sheet values and record timing
    t8 = time.time()
    try:
        noi_resp = sheets_service.spreadsheets().values().get(
            spreadsheetId=copied_sheet_id,
            range="'NOI'"
        ).execute()
        noi_values = noi_resp.get("values", [])
    except Exception as e:
        print(f"⚠️ Failed to fetch NOI sheet values: {e}")
        noi_values = []
    t9 = time.time()
    timings['noi_fetch'] = t9 - t8

    print("timings", timings)

    return {
        "levered_irr": variables.get("Levered IRR"),
        "levered_moic": variables.get("Levered MOIC"),
        "variables": variables,
        "NOI": noi_values,
    }



def add_blank_row_and_column_to_sheets(spreadsheet, sheet_names):
    """
    For each sheet in sheet_names, adds a blank row at the top (row 1)
    and a blank column at the left (column A).
    If a sheet does not exist, it skips that sheet.
    """
    requests = []
    for sheet_name in sheet_names:
        try:
            ws = spreadsheet.worksheet(sheet_name)
            sheet_id = ws._properties["sheetId"]
        except Exception as e:
            print(f"⚠️ Sheet '{sheet_name}' not found, skipping. Error: {e}")
            continue
        # Insert blank row at the top
        requests.append({
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "inheritFromBefore": False
            }
        })
        # Insert blank column at the left
        requests.append({
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "inheritFromBefore": False
            }
        })
    if requests:
        spreadsheet.batch_update({"requests": requests})


def update_google_sheet_and_get_values_final(
    copied_sheet_id,
    copied_sheet_url,
    mapped_values,
    market_json,
    rental_assumptions_json,
    rental_growth_json,
    amenity_income_json,
    expenses_json,
    retail_income_json
):
    print("📤 Starting Google Sheet generation workflow...")
    timings = {}
    t0 = time.time()
    sheets_service = build("sheets", "v4", credentials=creds)

    spreadsheet = gs_client.open_by_key(copied_sheet_id)
    add_blank_row_and_column_to_sheets(spreadsheet, ["Amenity Income", 
                                                     "Operating Expenses", 
                                                     "Closing Costs", 
                                                     "Legal and Pre-Development Costs", 
                                                     "Reserves",
                                                     "Hard Costs",
                                                     "Rental Assumptions",
                                                     "Market Rent Assumptions"])

    print("📊 Fetching mapping sheets...")
    result = sheets_service.spreadsheets().values().batchGet(
        spreadsheetId=copied_sheet_id,
        ranges=[
            "'Table Mapping'",
            "'Variable Mapping'"
        ],
        valueRenderOption='FORMULA'
    ).execute()
    print("RESULT", result)
    table_mapping_values = result['valueRanges'][0].get('values', [])
    variable_mapping_values = result['valueRanges'][1].get('values', [])

    table_mapping_data = [dict(zip(table_mapping_values[0], row)) for row in table_mapping_values[1:]]
    variable_data = [dict(zip(variable_mapping_values[0], row)) for row in variable_mapping_values[1:]]

    # Step 3: Extract tables
    t4 = time.time()
    tables = extract_tables_for_storage_batch(copied_sheet_id, table_mapping_data, sheets_service)
    t5 = time.time()
    timings['extract_tables'] = t5 - t4
    print(f"✅ Extracted {len(tables)} tables in {timings['extract_tables']:.3f}s")

    # Step 4: Extract variables
    t6 = time.time()
    variables = extract_variables_from_sheet_batch(copied_sheet_id, variable_data, sheets_service)
    t7 = time.time()
    timings['extract_variables'] = t7 - t6
    print(f"📈 Extracted variables: {variables} in {timings['extract_variables']:.3f}s")

    # Total time
    timings['total'] = t7 - t0
    print(f"⏱️ Total time: {timings['total']:.3f}s")
    print("🎯 Finished sheet generation and data extraction.")
    print(f"⏱️ Timings: {timings}")

    return {
        "sheet_url": copied_sheet_url,
        "levered_irr": variables.get("Levered IRR"),
        "levered_moic": variables.get("Levered MOIC"),
        "variables": variables,
        "tables": tables,
        "timings": timings
    }

def insert_expense_rows_to_sheet(spreadsheet, sheet_name, expenses):
    """
    Inserts rows into the given expense sheet based on the provided (already filtered) expenses.
    Always inserts the rows *between* the header row (1) and Total row (2).
    Includes dynamic formulas in columns J (Month 0) to EL (Month 132).
    """
    ws = spreadsheet.worksheet(sheet_name)
    sheet_id = ws._properties["sheetId"]

    from string import ascii_uppercase

    def generate_excel_columns(start='J', end='EL'):
        def col_index(c):
            result = 0
            for i, char in enumerate(reversed(c)):
                result += (ascii_uppercase.index(char) + 1) * (26 ** i)
            return result

        def col_name(index):
            name = ""
            while index > 0:
                index, remainder = divmod(index - 1, 26)
                name = chr(65 + remainder) + name
            return name

        start_idx = col_index(start)
        end_idx = col_index(end)
        return [col_name(i) for i in range(start_idx, end_idx + 1)]

    # Example usage
    month_columns = generate_excel_columns('J', 'EL')
    print(month_columns[:5], "...", month_columns[-5:])

    # Step 1: Insert rows between row 1 and 2 (no inheritFromBefore)
    insert_request = {
        "requests": [{
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 1,
                    "endIndex": 1 + len(expenses)
                },
                "inheritFromBefore": False
            }
        }]
    }
    spreadsheet.batch_update(insert_request)

    # Step 2: Separate out "Total percent of other expenses"
    percent_expense = next((e for e in expenses if e["factor"] == "Total percent of other expenses"), None)
    normal_expenses = [e for e in expenses if e["factor"] != "Total percent of other expenses"]

    rows_to_insert = []

    # Add normal expenses first
    for i, exp in enumerate(normal_expenses):
        row_index = i + 2  # 1-based index after header row

        name = exp.get("name", "")
        factor = exp.get("factor", "")
        cost_per = float(exp.get("cost_per") or 0)
        start_month = int(exp.get("start_month") or 0)
        end_month = int(exp.get("end_month") or 0)

        # Handle statistic
        statistic = "" if factor == "Total" else float(exp.get("statistic") or 0)

        # Total formula logic
        if factor == "Total":
            total_formula = f"=C{row_index}"
        elif factor in {"per Unit", "per SF", "per Month"}:
            total_formula = f"=C{row_index}*D{row_index}"
        else:
            total_formula = ""

        row = [name, factor, cost_per, statistic, total_formula, "", start_month, end_month, ""]

        # Monthly formulas from J to EL
        for month_offset, col_letter in enumerate(month_columns):
            formula = f"=IF(AND($G{row_index}={col_letter}$1,$H{row_index}={col_letter}$1),$E{row_index},IF($G{row_index}>{col_letter}$1,0,IF($H{row_index}<{col_letter}$1,0,$E{row_index}/($H{row_index}-$G{row_index}+1))))"
            row.append(formula)

        rows_to_insert.append(row)

    # Add "Total percent of other expenses" as the last row (if it exists)
    if percent_expense:
        percent_row_index = len(rows_to_insert) + 2  # Inserted after the rest
        name = percent_expense.get("name", "")
        cost_per_val = float(percent_expense.get("cost_per") or 0)
        cost_per = f"{cost_per_val}%"  # Add percent sign

        start_month = int(percent_expense.get("start_month") or 0)
        end_month = int(percent_expense.get("end_month") or 0)

        # Statistic = sum of previous total column; if no prior rows, use 0 to avoid SUM(E2:E1)
        statistic_formula = "0" if len(rows_to_insert) == 0 else f"=SUM(E2:E{percent_row_index - 1})"
        total_formula = f"=C{percent_row_index}*D{percent_row_index}"

        row = [name, "Total percent of other expenses", cost_per, statistic_formula, total_formula, "", start_month, end_month, ""]

        for col_letter in month_columns:
            formula = f"=IF(AND($G{percent_row_index}={col_letter}$1,$H{percent_row_index}={col_letter}$1),$E{percent_row_index},IF($G{percent_row_index}>{col_letter}$1,0,IF($H{percent_row_index}<{col_letter}$1,0,$E{percent_row_index}/($H{percent_row_index}-$G{percent_row_index}+1))))"
            row.append(formula)

        rows_to_insert.append(row)

    # Step 3: Update the inserted rows in the sheet
    start_row = 2
    end_row = 1 + len(rows_to_insert)
    end_col_letter = "EL"
    range_str = f"'{sheet_name}'!A{start_row}:{end_col_letter}{end_row}"
    spreadsheet.values_update(
        range_str,
        params={"valueInputOption": "USER_ENTERED"},
        body={"range": range_str, "values": rows_to_insert}
    )

    # Step 4: Update the Total row (now at row end_row + 1)
        # Step 4: Update the Total row (now at row end_row + 1)
    total_row_index = end_row + 1

    # Update E (total column)
    total_formula_value = f"=SUM(E2:E{end_row})"
    spreadsheet.values_update(
        f"{sheet_name}!E{total_row_index}",
        params={"valueInputOption": "USER_ENTERED"},
        body={"range": f"{sheet_name}!E{total_row_index}", "values": [[total_formula_value]]}
    )

    # Update each month column (J to EL) with a SUM formula
    month_sums = []
    for col_letter in month_columns:
        formula = f"=SUM({col_letter}2:{col_letter}{end_row})"
        month_sums.append(formula)

    # Write the full totals row (starting from column J)
    spreadsheet.values_update(
        f"{sheet_name}!J{total_row_index}:{month_columns[-1]}{total_row_index}",
        params={"valueInputOption": "USER_ENTERED"},
        body={"range": f"{sheet_name}!J{total_row_index}:{month_columns[-1]}{total_row_index}", "values": [month_sums]}
    )

    print(f"✅ Inserted {len(rows_to_insert)} expense rows into '{sheet_name}' and updated total row.")

def update_user_model_expense_table(copied_sheet_id, sheet_name, expenses):
    spreadsheet = gs_client.open_by_key(copied_sheet_id)
    
    # Clear existing rows between header and total row
    clear_expense_table_rows(spreadsheet, sheet_name)

    # Load model_variable_mapping from the sheet so we can reference mapped cells
    sheets_service = build("sheets", "v4", credentials=creds)
    mapping_result = sheets_service.spreadsheets().values().batchGet(
        spreadsheetId=copied_sheet_id,
        ranges=[
            "'Model Variable Mapping'",
        ],
        valueRenderOption='FORMULA'
    ).execute()

    model_mapping_values = mapping_result['valueRanges'][0].get('values', [])
    if not model_mapping_values:
        raise Exception("Model Variable Mapping sheet is empty or missing")

    # Convert to DataFrame for model_variable_mapping, pad rows to header length
    header = model_mapping_values[0]
    normalized_rows = [row + [''] * (len(header) - len(row)) for row in model_mapping_values[1:]]
    df = pd.DataFrame(normalized_rows, columns=header)

    # Build payloads and apply
    insert_request, update_payloads, format_requests = insert_expense_rows_to_sheet_payloads(
        spreadsheet, sheet_name, expenses, df
    )

    # Perform row inserts
    spreadsheet.batch_update({"requests": [insert_request, *format_requests]})
    # Apply values/formulas
    spreadsheet.values_batch_update({"valueInputOption": "USER_ENTERED", "data": update_payloads})

    # spreadsheet.batch_update({"requests": format_requests})
    

    print(f"✅ Updated expense table for '{sheet_name}' with {len(expenses)} rows using payloads and number formats.")

def clear_expense_table_rows(spreadsheet, sheet_name):
    """
    Clears all rows in each given sheet from the header row through the row containing
    'Total <Sheet Name>' in column A. Leaves only the header row and the total row.
    """

    try:
        ws = spreadsheet.worksheet(sheet_name)
        sheet_id = ws._properties['sheetId']
        all_values = ws.get_all_values()

        # Find row with "Total <sheet_name>" in column A
        target_label = f"Total {sheet_name}"
        target_row_index = None
        for i, row in enumerate(all_values, start=1):
            if row and row[0].strip() == target_label:
                target_row_index = i
                break

        if target_row_index is None:
            print(f"⚠️ Could not find 'Total {sheet_name}' in column A")
            return

        # Rows to clear: everything from row 2 up to (but not including) the total row
        if target_row_index <= 2:
            print(f"ℹ️ Nothing to clear in '{sheet_name}'")
            return

        # Build delete request
        delete_request = {
            "deleteDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 1,  # Row 2 (0-indexed)
                    "endIndex": target_row_index - 1  # Up to row before 'Total'
                }
            }
        }

        spreadsheet.batch_update({"requests": [delete_request]})
        print(f"✅ Cleared rows 2 to {target_row_index - 1} in '{sheet_name}'")

    except Exception as e:
        print(f"❌ Error processing sheet '{sheet_name}': {e}")

def clean_number(value):
    """Clean and convert string values to numbers"""
    if isinstance(value, str):
        s = value.strip().replace("'", "").replace(",", "")
        # Handle percent values like '33.5%'
        if s.endswith('%'):
            try:
                return float(s.strip('%'))
            except ValueError:
                return s
        # Handle MOIC values like '2.95x'
        if s.lower().endswith('x'):
            try:
                return float(s[:-1])
            except ValueError:
                return s
        try:
            return float(s)
        except ValueError:
            return value
    return value

def update_inputs(worksheet, row_val, col_val, purchase_price_cell="E62", exit_cap_cell="G20"):
    """Update purchase price and exit cap rate cells in the worksheet"""
    
    # Extract cell references from full locations (remove sheet name if present)
    def extract_cell_ref(location):
        if '!' in location:
            return location.split('!')[1]
        return location
    
    purchase_price_ref = extract_cell_ref(purchase_price_cell)
    exit_cap_ref = extract_cell_ref(exit_cap_cell)
    
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{worksheet.spreadsheet.id}/values:batchUpdate"
    body = {
        "valueInputOption": "RAW",
        "data": [
            {"range": f"{worksheet.title}!{purchase_price_ref}", "values": [[clean_number(row_val)]]},
            {"range": f"{worksheet.title}!{exit_cap_ref}", "values": [[f"{clean_number(col_val)}%"]]},
        ],
    }
    response = worksheet.spreadsheet.client.session.request("POST", url=url, json=body)
    response.raise_for_status()

def get_range_string(corner, num_rows, num_cols):
    """Generate range string from corner cell and dimensions"""
    col_letter = corner[0]
    row_number = int(corner[1:])
    end_col = chr(ord(col_letter) + num_cols - 1)
    end_row = row_number + num_rows - 1
    return f"{corner}:{end_col}{end_row}"

def generate_sensitivity_analysis_tables(sheet_id, max_price, min_cap_rate):
    """
    Generate sensitivity analysis tables for IRR and MOIC based on purchase price and exit cap rate variations
    """
    print(f"🔧 [DEBUG] Starting sensitivity analysis for sheet_id: {sheet_id}")
    print(f"💰 [DEBUG] Max price: {max_price}, Min cap rate: {min_cap_rate}")
    
    # Configure Google Sheets API
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    
    if SERVICE_ACCOUNT_FILE and os.path.exists("./" + SERVICE_ACCOUNT_FILE):
        print("🔑 Using local service account credentials for sensitivity analysis")
        creds = Credentials.from_service_account_file("./" + SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    else:
        print("🔐 Using default GCP credentials for sensitivity analysis")
        creds, _ = default(scopes=SCOPES)
    
    client = gspread.authorize(creds)
    sheets_service = build("sheets", "v4", credentials=creds)
    
    try:
        # Extract model_variable_mapping from the Google Sheet
        print("📊 [DEBUG] Extracting model variable mapping...")
        result = sheets_service.spreadsheets().values().batchGet(
            spreadsheetId=sheet_id,
            ranges=[
                "'Model Variable Mapping'"
            ],
            valueRenderOption='FORMULA'
        ).execute()
        
        model_mapping_values = result['valueRanges'][0].get('values', [])
        # For all rows with length 3, append two empty strings
        model_mapping_values = [row + ['', ''] if len(row) == 3 else row for row in model_mapping_values]
        
        # Convert to DataFrame for model_variable_mapping
        df = pd.DataFrame(model_mapping_values[1:], columns=model_mapping_values[0])
        # print(f"✅ [DEBUG] Extracted model_variable_mapping with {len(df)} entries")
        
        # # Debug: Print all available entries to see what we have
        # print("🔍 [DEBUG] Available entries in model variable mapping:")
        for _, row in df.iterrows():
            section = row.get('section', '')
            field_key = row.get('field_key', '')
            location = row.get('location', '')
            print(f"  - Section: '{section}', Field: '{field_key}', Location: '{location}'")
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        assumptions_ws = spreadsheet.worksheet("Assumptions")
        
        # print(f"✅ [DEBUG] Opened spreadsheet and Assumptions worksheet")
        
        # Get cell locations using model_variable_mapping
        purchase_price_cell = get_mapped_cell_location(df, 'General Property Assumptions', 'Acquisition Price')

        
        exit_cap_cell = get_mapped_cell_location(df, 'Exit Assumptions', 'Multifamily Applied Exit Cap Rate')
        if not exit_cap_cell:
            exit_cap_cell = get_mapped_cell_location(df, 'Exit Assumptions', 'Retail Applied Exit Cap Rate')


        IRR_cell = get_mapped_cell_location(df, 'Other Reference', 'Levered IRR')
        MOIC_cell = get_mapped_cell_location(df, 'Other Reference', 'Levered MOIC')
        
        # Get sensitivity reference cell locations
        sensitivity_purchase_price_1_cell = get_mapped_cell_location(df, 'Other Reference', 'Sensitivity Purchase Price 1')
        sensitivity_purchase_price_2_cell = get_mapped_cell_location(df, 'Other Reference', 'Sensitivity Purchase Price 2')
        sensitivity_exit_cap_rate_1_cell = get_mapped_cell_location(df, 'Other Reference', 'Sensitivity Exit Cap Rate 1')
        sensitivity_exit_cap_rate_2_cell = get_mapped_cell_location(df, 'Other Reference', 'Sensitivity Exit Cap Rate 2')
        
        # Construct table start locations by combining column from purchase price and row from exit cap rate
        # IRR table: column from purchase_price_1, row from exit_cap_rate_1
        if sensitivity_purchase_price_1_cell and sensitivity_exit_cap_rate_1_cell:
            IRR_left_corner = f"{sensitivity_purchase_price_1_cell[0]}{sensitivity_exit_cap_rate_1_cell[1:]}"
        else:
            IRR_left_corner = None
            
        # MOIC table: column from purchase_price_2, row from exit_cap_rate_2  
        if sensitivity_purchase_price_2_cell and sensitivity_exit_cap_rate_2_cell:
            MOIC_left_corner = f"{sensitivity_purchase_price_2_cell[0]}{sensitivity_exit_cap_rate_2_cell[1:]}"
        else:
            MOIC_left_corner = None
        
        # Validate that all required cells were found
        required_cells = {
            'purchase_price_cell': purchase_price_cell,
            'exit_cap_cell': exit_cap_cell,
            'IRR_cell': IRR_cell,
            'MOIC_cell': MOIC_cell,
            'IRR_left_corner': IRR_left_corner,
            'MOIC_left_corner': MOIC_left_corner,
            'sensitivity_purchase_price_1_cell': sensitivity_purchase_price_1_cell,
            'sensitivity_purchase_price_2_cell': sensitivity_purchase_price_2_cell,
            'sensitivity_exit_cap_rate_1_cell': sensitivity_exit_cap_rate_1_cell,
            'sensitivity_exit_cap_rate_2_cell': sensitivity_exit_cap_rate_2_cell
        }
        
        missing_cells = [name for name, location in required_cells.items() if not location]
        if missing_cells:
            raise Exception(f"Missing cell locations for: {', '.join(missing_cells)}")
        
        # print(f"📊 [DEBUG] Cell locations from model_variable_mapping:")
        for name, location in required_cells.items():
            print(f"  - {name}: {location}")
        
        num_rows = 5
        num_cols = 5
        
        # print(f"📊 [DEBUG] Will update sensitivity reference cells:")
        # print(f"  - Purchase Price 1: {sensitivity_purchase_price_1_cell}")
        # print(f"  - Purchase Price 2: {sensitivity_purchase_price_2_cell}")
        # print(f"  - Exit Cap Rate 1: {sensitivity_exit_cap_rate_1_cell}")
        # print(f"  - Exit Cap Rate 2: {sensitivity_exit_cap_rate_2_cell}")
        
        # Calculate the sensitivity ranges based on max_price and min_cap_rate
        # For purchase prices: range around max_price (±2% or similar)
        purchase_price_1 = max_price 
        purchase_price_2 = max_price 
        
        # For cap rates: range around min_cap_rate (±0.5% or similar)
        exit_cap_rate_1 = min_cap_rate
        exit_cap_rate_2 = min_cap_rate
        
        # print(f"📊 [DEBUG] Calculated sensitivity ranges:")
        # print(f"  - Purchase Price Range: {purchase_price_1:,.0f} to {purchase_price_2:,.0f}")
        # print(f"  - Cap Rate Range: {exit_cap_rate_1:.2f}% to {exit_cap_rate_2:.2f}%")
        
        # Update the sensitivity reference cells
        update_sensitivity_reference_cells(assumptions_ws, 
                                         purchase_price_1, purchase_price_2,
                                         exit_cap_rate_1, exit_cap_rate_2,
                                         sensitivity_purchase_price_1_cell, sensitivity_purchase_price_2_cell,
                                         sensitivity_exit_cap_rate_1_cell, sensitivity_exit_cap_rate_2_cell)
        

        time.sleep(0.1)
        
        # NOW read the actual header values from the sheet after they've been updated
        # This will get the static values, not formulas
        # print(f"📊 [DEBUG] Reading updated header values from the sheet...")
        

        def extract_cell_ref(location):
            if '!' in location:
                return location.split('!')[1]
            return location
        
        def build_vertical_range(start_a1: str, count: int) -> str:
            start_row, start_col = a1_to_rowcol(start_a1)
            end_row = start_row + count - 1
            start_cell = rowcol_to_a1(start_row, start_col)
            end_cell = rowcol_to_a1(end_row, start_col)
            return f"{start_cell}:{end_cell}"
        
        def build_horizontal_range(start_a1: str, count: int) -> str:
            start_row, start_col = a1_to_rowcol(start_a1)
            end_col = start_col + count - 1
            start_cell = rowcol_to_a1(start_row, start_col)
            end_cell = rowcol_to_a1(start_row, end_col)
            return f"{start_cell}:{end_cell}"
        
        exit1_ref = extract_cell_ref(sensitivity_exit_cap_rate_1_cell)
        price1_ref = extract_cell_ref(sensitivity_purchase_price_1_cell)
        exit2_ref = extract_cell_ref(sensitivity_exit_cap_rate_2_cell)
        price2_ref = extract_cell_ref(sensitivity_purchase_price_2_cell)
        
        row_header_range_1 = build_vertical_range(exit1_ref, 5)
        col_header_range_1 = build_horizontal_range(price1_ref, 5)
        row_header_range_2 = build_vertical_range(exit2_ref, 5)
        col_header_range_2 = build_horizontal_range(price2_ref, 5)
        
        # print(f"📐 [DEBUG] Header ranges (table 1): rows={row_header_range_1}, cols={col_header_range_1}")
        # print(f"📐 [DEBUG] Header ranges (table 2): rows={row_header_range_2}, cols={col_header_range_2}")
        
        # Read header values from both tables; use table 1 as canonical inputs
        row_inputs_raw_1 = assumptions_ws.get(row_header_range_1)
        col_inputs_raw_1 = assumptions_ws.get(col_header_range_1)
        row_inputs_raw_2 = assumptions_ws.get(row_header_range_2)
        col_inputs_raw_2 = assumptions_ws.get(col_header_range_2)
        
        # print(f"📋 [DEBUG] Raw row inputs (table 1): {row_inputs_raw_1}")
        # print(f"📋 [DEBUG] Raw col inputs (table 1): {col_inputs_raw_1}")
        # print(f"📋 [DEBUG] Raw row inputs (table 2): {row_inputs_raw_2}")
        # print(f"📋 [DEBUG] Raw col inputs (table 2): {col_inputs_raw_2}")
        
        # Use table 1 values as the driving inputs
        row_inputs_raw = row_inputs_raw_1
        col_inputs_raw = col_inputs_raw_1
        
        # Process row inputs (cap rates) - filter out empty values
        row_inputs = [r[0] for r in row_inputs_raw if r and len(r) > 0 and r[0] != '']
        # Process column inputs (purchase prices) - filter out empty values
        col_inputs = [c for c in (col_inputs_raw[0] if col_inputs_raw else []) if c != '']
        
        # print(f"📋 [DEBUG] Final row inputs (cap rates): {row_inputs}")
        # print(f"📋 [DEBUG] Final col inputs (purchase prices): {col_inputs}")
        # print(f"📋 [DEBUG] Number of rows: {len(row_inputs)}, Number of cols: {len(col_inputs)}")
        
        # Save original input values so we can restore at the end
        purchase_price_ref = extract_cell_ref(purchase_price_cell)
        exit_cap_ref = extract_cell_ref(exit_cap_cell)
        original_purchase_price = assumptions_ws.acell(purchase_price_ref).value
        original_exit_cap = assumptions_ws.acell(exit_cap_ref).value
        # print(f"💾 [DEBUG] Saved original values - {purchase_price_cell}: {original_purchase_price}, {exit_cap_cell}: {original_exit_cap}")
        
        irr_grid = []
        moic_grid = []
        irr_grid_raw = []
        moic_grid_raw = []
        
        # Generate sensitivity analysis grid
        for i, cap_rate in enumerate(row_inputs):
            # print(f"🔄 [DEBUG] Processing row {i+1}/{len(row_inputs)}: cap_rate={cap_rate}")
            irr_row = []
            moic_row = []
            irr_row_raw = []
            moic_row_raw = []
            for j, purchase_price in enumerate(col_inputs):
                # print(f"  🔄 [DEBUG] Processing cell {i+1},{j+1}: cap_rate={cap_rate}, purchase_price={purchase_price}")
                
                # Update inputs in the sheet - cap_rate goes to exit_cap_cell, purchase_price goes to purchase_price_cell
                # IMPORTANT: Only update the main input cells, NOT the header cells
                update_inputs(assumptions_ws, purchase_price, cap_rate, purchase_price_cell, exit_cap_cell)
                
                # Wait for calculations to complete
                import time as _time
                _time.sleep(0.1)
                
                # Get results from the IRR and MOIC cells (NOT from header cells)
                IRR_ref = extract_cell_ref(IRR_cell)
                MOIC_ref = extract_cell_ref(MOIC_cell)
                irr_result = assumptions_ws.acell(IRR_ref).value
                moic_result = assumptions_ws.acell(MOIC_ref).value
                
                # print(f"    📊 [DEBUG] Results - IRR: {irr_result}, MOIC: {moic_result}")
                
                irr_row.append(clean_number(irr_result))
                moic_row.append(clean_number(moic_result))
                irr_row_raw.append(irr_result)
                moic_row_raw.append(moic_result)
            
            irr_grid.append(irr_row)
            moic_grid.append(moic_row)
            irr_grid_raw.append(irr_row_raw)
            moic_grid_raw.append(moic_row_raw)
        
        # Write the full grids back to the sheet (table body only)
        # Compute table body top-left corners from dynamic header references
        row1, _col_from_exit = a1_to_rowcol(exit1_ref)
        _row_from_price, col1 = a1_to_rowcol(price1_ref)
        irr_corner = rowcol_to_a1(row1, col1)
        irr_range = get_range_string(irr_corner, len(row_inputs), len(col_inputs))

        row2, _ = a1_to_rowcol(exit2_ref)
        _r2, col2 = a1_to_rowcol(price2_ref)
        moic_corner = rowcol_to_a1(row2, col2)
        moic_range = get_range_string(moic_corner, len(row_inputs), len(col_inputs))

        # print(f"📝 [DEBUG] Writing IRR grid to {irr_range} and MOIC grid to {moic_range}")
        assumptions_ws.update(irr_range, irr_grid_raw)
        assumptions_ws.update(moic_range, moic_grid_raw)
        
        # Restore original inputs
        update_inputs(assumptions_ws, original_purchase_price, original_exit_cap, purchase_price_cell, exit_cap_cell)
        # print(f"✅ [DEBUG] Sensitivity analysis completed successfully")
        
        # Return the results with headers
        return {
            'irr_table': {
                'capRates': [clean_number(rate) for rate in row_inputs],
                'acquisitionPrices': [clean_number(price) for price in col_inputs],
                'values': [[val for val in row] for row in irr_grid]
            },
            'moic_table': {
                'capRates': [clean_number(rate) for rate in row_inputs],
                'acquisitionPrices': [clean_number(price) for price in col_inputs],
                'values': [[val for val in row] for row in moic_grid]
            }
        }
        
    except Exception as e:
        print(f"❌ [DEBUG] Error in sensitivity analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e

def update_sensitivity_reference_cells(worksheet, purchase_price_1, purchase_price_2, exit_cap_rate_1, exit_cap_rate_2,
                                     price_1_cell, price_2_cell, cap_rate_1_cell, cap_rate_2_cell):
    """Update the sensitivity reference cells with the calculated ranges"""
    
    # Extract cell references from full locations (remove sheet name if present)
    def extract_cell_ref(location):
        if '!' in location:
            return location.split('!')[1]
        return location
    
    price_1_ref = extract_cell_ref(price_1_cell)
    price_2_ref = extract_cell_ref(price_2_cell)
    cap_rate_1_ref = extract_cell_ref(cap_rate_1_cell)
    cap_rate_2_ref = extract_cell_ref(cap_rate_2_cell)
    
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{worksheet.spreadsheet.id}/values:batchUpdate"
    body = {
        "valueInputOption": "RAW",
        "data": [
            {"range": f"{worksheet.title}!{price_1_ref}", "values": [[purchase_price_1]]},
            {"range": f"{worksheet.title}!{price_2_ref}", "values": [[purchase_price_2]]},
            {"range": f"{worksheet.title}!{cap_rate_1_ref}", "values": [[f"{exit_cap_rate_1}%"]]},
            {"range": f"{worksheet.title}!{cap_rate_2_ref}", "values": [[f"{exit_cap_rate_2}%"]]},
        ],
    }
    response = worksheet.spreadsheet.client.session.request("POST", url=url, json=body)
    response.raise_for_status()
    print(f"✅ [DEBUG] Updated sensitivity reference cells")

def get_rent_roll_vacancy_formula_updates(rent_roll_ws, rental_assumptions_json, rental_growth_json,
                                          base_start_row=14, start_col_index=10, num_columns=131, units_start_row=7):
    """
    Write vacancy calculation formulas to the Rent Roll sheet.
    Start row = base_start_row + len(rental growth rows (type=='rental'))
    One row per rental unit (len(rental_assumptions_json)).
    Formula pattern per cell (for row r, column letter L):
      =IFERROR(IF($Dr=0,$Gr,IF(L$9<$Er,$Gr,IF(L$9>=$Fr,$Hr,$Gr)))*INDEX(L$3:L$5,MATCH($Cr,$B$3:$B$5,0)),0)
    """
    rental_growth_rows = len([entry for entry in rental_growth_json if entry.get("type") == "rental"])
    start_row = base_start_row + rental_growth_rows
    units_start_row = units_start_row + rental_growth_rows
    num_rows = len(rental_assumptions_json)

    formulas = []
    for i in range(num_rows):
        row_num = start_row + i  # vacancy output row
        unit_row = units_start_row + i  # corresponding unit data row (e.g., 10,11,...)
        row_formulas = []
        for j in range(num_columns):
            col_index = start_col_index + j
            col_letter = rowcol_to_a1(1, col_index).replace("1", "")
            # Build the vacancy formula for this cell using dynamic growth range (rows 3 .. 2 + rental_growth_rows)
            end_row = 2 + rental_growth_rows  # e.g., 4 when 2 rates, 5 when 3 rates
            formula = (
                f"=IFERROR(IF($D{unit_row}=0,$G{unit_row},IF({col_letter}$9<$E{unit_row},$G{unit_row},"
                f"IF({col_letter}$9>=$F{unit_row},$H{unit_row},$G{unit_row})))*"
                f"INDEX({col_letter}$3:{col_letter}${end_row},MATCH($C{unit_row},$B$3:$B${end_row},0)),0)"
            )
            row_formulas.append(formula)
        formulas.append(row_formulas)

    start_col_letter = rowcol_to_a1(1, start_col_index).replace("1", "")
    end_col_index = start_col_index + num_columns - 1
    end_col_letter = rowcol_to_a1(1, end_col_index).replace("1", "")

    value_payload = {
        "range": f"'Rent Roll Model'!{start_col_letter}{start_row}:{end_col_letter}{start_row + num_rows - 1}",
        "values": formulas
    }

    format_request = {
        "repeatCell": {
            "range": {
                "sheetId": rent_roll_ws._properties["sheetId"],
                "startRowIndex": start_row - 1,
                "endRowIndex": start_row - 1 + num_rows,
                "startColumnIndex": start_col_index - 1,
                "endColumnIndex": end_col_index
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "NUMBER",
                        "pattern": "#,##0"
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat"
        }
    }

    return value_payload, format_request

def get_rent_roll_vacancy_row_inserts_and_updates(
    rent_roll_ws,
    rental_assumptions_json,
    rental_growth_json,
    base_start_row=14,
    start_col_index=10,
    num_columns=131,
    units_start_row=7
):
    # rows to insert begin at 14 + growth_len + num_units
    growth_len = len([entry for entry in rental_growth_json if entry.get("type") == "rental"])
    num_units = len(rental_assumptions_json)
    start_row = base_start_row + growth_len + num_units

    insert_request = {
        "insertDimension": {
            "range": {
                "sheetId": rent_roll_ws._properties["sheetId"],
                "dimension": "ROWS",
                "startIndex": start_row - 1,
                "endIndex": start_row - 1 + num_units
            },
            "inheritFromBefore": False
        }
    }

    # Generate formulas targeted to this block (pass base_start_row plus units to offset inside helper)
    value_payload, format_request = get_rent_roll_vacancy_formula_updates(
        rent_roll_ws=rent_roll_ws,
        rental_assumptions_json=rental_assumptions_json,
        rental_growth_json=rental_growth_json,
        base_start_row=base_start_row + num_units,
        start_col_index=start_col_index,
        num_columns=num_columns,
        units_start_row=units_start_row
    )

    return insert_request, value_payload, format_request

def get_rent_roll_vacancy_sum_row_update(
    rent_roll_ws,
    rental_assumptions_json,
    rental_growth_json,
    base_start_row=14,
    start_col_index=10,
    num_columns=131
):
    # Compute where the vacancy block starts and ends
    rental_growth_rows = len([entry for entry in rental_growth_json if entry.get("type") == "rental"])
    num_units = len(rental_assumptions_json)
    vacancy_start_row = base_start_row + rental_growth_rows + num_units  # first vacancy row
    sum_row_index = vacancy_start_row + num_units  # row immediately below vacancy block

    # Build SUM formulas across the month columns (J..)
    month_formulas = []
    for j in range(num_columns):
        col_index = start_col_index + j
        col_letter = rowcol_to_a1(1, col_index).replace("1", "")
        month_formulas.append(
            f"=SUM({col_letter}{vacancy_start_row}:{col_letter}{vacancy_start_row + num_units - 1})"
        )

    start_col_letter = rowcol_to_a1(1, start_col_index).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col_index + num_columns - 1).replace("1", "")

    update_payload = {
        "range": f"'Rent Roll Model'!{start_col_letter}{sum_row_index}:{end_col_letter}{sum_row_index}",
        "values": [month_formulas]
    }

    return update_payload

def get_noi_walk_egi_row_update_payload(
    amenity_income_json,
    sheet_name="NOI Walk",
    start_col=5,        # Column E
    num_months=132,
    amenity_start_row=25,
    target_base_row=27
):
    """
    Build values batch update payload for the NOI Walk sheet.
    Writes formulas on row (target_base_row + len(amenity_income_json)).
    For each month column L, the formula is:
      =SUM(L25:L<amenity_end>,L22) + IF(L15>=$H$5, L23, 0)
    Where L is the dynamic month column letter.
    """
    amenity_len = len(amenity_income_json)
    target_row = target_base_row + amenity_len
    amenity_end_row = amenity_start_row + amenity_len - 1

    row_formulas = []
    for j in range(num_months):
        col_index = start_col + j
        col_letter = rowcol_to_a1(1, col_index).replace("1", "")
        formula = (
            f"=SUM({col_letter}{amenity_start_row}:{col_letter}{amenity_end_row + 2},{col_letter}22)"
            f"+IF({col_letter}15>=$H$5,{col_letter}23,0)"
        )
        row_formulas.append(formula)

    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col + num_months - 1).replace("1", "")

    return {
        "range": f"'{sheet_name}'!{start_col_letter}{target_row}:{end_col_letter}{target_row}",
        "values": [row_formulas]
    }


def get_noi_walk_egi_row_update_payload_industrial(
    amenity_income_json,
    sheet_name="NOI Walk",
    start_col=5,        # Column E
    num_months=132,
    amenity_start_row=25,
    target_base_row=27
):
    """
    Build values batch update payload for the NOI Walk sheet.
    Writes formulas on row (target_base_row + len(amenity_income_json)).
    For each month column L, the formula is:
      =SUM(L25:L<amenity_end>,L22) + IF(L15>=$H$5, L23, 0)
    Where L is the dynamic month column letter.
    """
    amenity_len = len(amenity_income_json)
    target_row = target_base_row + amenity_len
    amenity_end_row = amenity_start_row + amenity_len - 1

    row_formulas = []
    for j in range(num_months):
        col_index = start_col + j
        col_letter = rowcol_to_a1(1, col_index).replace("1", "")
        formula = (
            f"=SUM({col_letter}{amenity_start_row}:{col_letter}{amenity_end_row},{col_letter}{amenity_start_row -3})"
        )
        row_formulas.append(formula)

    start_col_letter = rowcol_to_a1(1, start_col).replace("1", "")
    end_col_letter = rowcol_to_a1(1, start_col + num_months - 1).replace("1", "")

    return {
        "range": f"'{sheet_name}'!{start_col_letter}{target_row}:{end_col_letter}{target_row}",
        "values": [row_formulas]
    }





























































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































